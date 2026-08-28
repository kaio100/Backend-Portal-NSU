from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[3]
SOURCE = ROOT / "relatorios" / "Apresentacao_Resultados_Portal_NFSe_Online.xlsx"
OUTPUT = ROOT / "relatorios" / "Apresentacao_Resultados_Portal_NFSe_Executiva.xlsx"

NAVY = "0B1F3A"
BLUE = "2563EB"
CYAN = "0891B2"
GREEN = "059669"
AMBER = "D97706"
RED = "DC2626"
WHITE = "FFFFFF"
TEXT = "172033"
MUTED = "64748B"
LIGHT = "F4F7FB"
GRID = "DDE5F0"


def card(ws, start_col: int, start_row: int, label: str, value, explanation: str, color: str, fmt: str = "#,##0"):
    end_col = start_col + 2
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=end_col)
    ws.merge_cells(start_row=start_row + 3, start_column=start_col, end_row=start_row + 4, end_column=end_col)
    label_cell = ws.cell(start_row, start_col, label)
    value_cell = ws.cell(start_row + 1, start_col, value)
    note_cell = ws.cell(start_row + 3, start_col, explanation)
    label_cell.fill = PatternFill("solid", fgColor=color)
    label_cell.font = Font(size=11, bold=True, color=WHITE)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.fill = PatternFill("solid", fgColor=WHITE)
    value_cell.font = Font(size=25, bold=True, color=TEXT)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.number_format = fmt
    note_cell.fill = PatternFill("solid", fgColor=LIGHT)
    note_cell.font = Font(size=9, color=MUTED)
    note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(color=GRID), right=Side(color=GRID), top=Side(color=GRID), bottom=Side(color=GRID))
    for row in range(start_row, start_row + 5):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = border


def section(ws, row: int, title: str, text: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 2, end_column=12)
    ws.cell(row, 1, title)
    ws.cell(row, 1).font = Font(size=17, bold=True, color=TEXT)
    ws.cell(row + 1, 1, text)
    ws.cell(row + 1, 1).font = Font(size=11, color=MUTED)
    ws.cell(row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")


def main():
    wb = load_workbook(SOURCE, data_only=False)
    if "Apresentação" in wb.sheetnames:
        del wb["Apresentação"]
    ws = wb.create_sheet("Apresentação", 0)
    base = wb["Base de Notas"]
    detail = wb["Dashboard Executivo"]

    total_notes = int(detail["A6"].value or 0)
    total_value = float(detail["C6"].value or 0)
    companies_with_notes = int(detail["E6"].value or 0)
    pdfs = int(detail["G6"].value or 0)
    users = int(detail["A10"].value or 0)
    accesses = int(detail["C10"].value or 0)
    active_certs = int(detail["E10"].value or 0)
    running_certs = int(detail["G10"].value or 0)
    files = int(detail["I10"].value or 0)
    notes_minute = float(detail["K10"].value or 0)

    monthly = defaultdict(lambda: {"total": 0, "tomadas": 0, "prestadas": 0})
    directions = Counter()
    statuses = Counter()
    alerts = Counter()
    for values in base.iter_rows(min_row=2, values_only=True):
        month = values[1] or "Sem data"
        direction = values[5] or "Não informado"
        status = values[9] or "Não informado"
        alert = values[13] or "Não"
        monthly[month]["total"] += 1
        monthly[month]["tomadas" if direction == "Tomado" else "prestadas"] += 1
        directions[direction] += 1
        statuses[status] += 1
        alerts[alert] += 1

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 75
    ws.freeze_panes = "A6"
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 13
    for row in range(1, 95):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 34

    ws.merge_cells("A1:L2")
    ws["A1"] = "O que o Portal NFS-e entregou"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(size=28, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A3:L4")
    ws["A3"] = "Uma visão simples do volume processado, da adoção pela equipe e do ganho operacional alcançado com a automação. Dados extraídos diretamente do ambiente de produção."
    ws["A3"].font = Font(size=12, color=MUTED)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    section(ws, 6, "1. Resultado em uma frase", "O sistema centralizou, organizou e disponibilizou milhares de notas e documentos, reduzindo uma rotina manual repetitiva para um processo automatizado e rastreável.")
    card(ws, 1, 10, "NOTAS PROCESSADAS", total_notes, "NFS-e disponíveis para consulta e conferência", BLUE)
    card(ws, 4, 10, "ARQUIVOS ORGANIZADOS", files, "XMLs, PDFs e demais documentos registrados", CYAN)
    card(ws, 7, 10, "EMPRESAS ATENDIDAS", companies_with_notes, "Empresas que já possuem notas processadas", GREEN)
    card(ws, 10, 10, "VALOR MOVIMENTADO", total_value, "Soma do valor dos serviços nas notas", AMBER, 'R$ #,##0')

    section(ws, 17, "2. Uso e operação", "Estes números mostram o alcance atual da solução e a capacidade disponível para continuar crescendo.")
    card(ws, 1, 21, "USUÁRIOS", users, "Pessoas cadastradas no portal", BLUE)
    card(ws, 4, 21, "ACESSOS", accesses, "Entradas registradas desde a ativação do controle", CYAN)
    card(ws, 7, 21, "CERTIFICADOS ATIVOS", active_certs, "Certificados aptos para consultar notas", GREEN)
    card(ws, 10, 21, "PDFs DISPONÍVEIS", pdfs, "Documentos prontos para consulta e download", AMBER)

    section(ws, 28, "3. Como o volume evoluiu", "O gráfico mostra quantas notas foram incorporadas ao portal em cada mês de emissão. Ele pode ser redimensionado livremente no Excel.")

    helper_col = 20
    headers = ["Mês", "Notas", "Tomadas", "Prestadas"]
    for idx, value in enumerate(headers, helper_col): ws.cell(1, idx, value)
    month_rows = [(m, v["total"], v["tomadas"], v["prestadas"]) for m, v in sorted(monthly.items()) if m != "Sem data"]
    for ridx, values in enumerate(month_rows, 2):
        for cidx, value in enumerate(values, helper_col): ws.cell(ridx, cidx, value)

    line = LineChart()
    line.title = "Notas incorporadas ao portal por mês"
    line.style = 13
    line.height, line.width = 10, 24
    line.visible_cells_only = False
    line.y_axis.title = "Quantidade de notas"
    line.x_axis.title = "Mês de emissão"
    line.add_data(Reference(ws, min_col=helper_col + 1, min_row=1, max_row=1 + len(month_rows)), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=helper_col, min_row=2, max_row=1 + len(month_rows)))
    ws.add_chart(line, "A32")

    section(ws, 52, "4. O que está sendo processado", "A separação entre serviços tomados e prestados ajuda a entender de onde vem o volume operacional.")
    dir_start = 2
    ws.cell(dir_start, helper_col + 5, "Tipo")
    ws.cell(dir_start, helper_col + 6, "Quantidade")
    for idx, (label, value) in enumerate(directions.most_common(), dir_start + 1):
        ws.cell(idx, helper_col + 5, label)
        ws.cell(idx, helper_col + 6, value)
    doughnut = DoughnutChart()
    doughnut.title = "Notas tomadas x prestadas"
    doughnut.height, doughnut.width = 9, 12
    doughnut.visible_cells_only = False
    doughnut.add_data(Reference(ws, min_col=helper_col + 6, min_row=dir_start, max_row=dir_start + len(directions)), titles_from_data=True)
    doughnut.set_categories(Reference(ws, min_col=helper_col + 5, min_row=dir_start + 1, max_row=dir_start + len(directions)))
    doughnut.dataLabels = DataLabelList()
    doughnut.dataLabels.showPercent = True
    doughnut.dataLabels.showLeaderLines = True
    ws.add_chart(doughnut, "A56")

    top_status = statuses.most_common(6)
    status_start = 2
    ws.cell(status_start, helper_col + 8, "Status")
    ws.cell(status_start, helper_col + 9, "Quantidade")
    for idx, (label, value) in enumerate(top_status, status_start + 1):
        ws.cell(idx, helper_col + 8, label)
        ws.cell(idx, helper_col + 9, value)
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Situação atual das notas"
    bar.height, bar.width = 9, 12
    bar.visible_cells_only = False
    bar.y_axis.title = "Status"
    bar.x_axis.title = "Quantidade"
    bar.add_data(Reference(ws, min_col=helper_col + 9, min_row=status_start, max_row=status_start + len(top_status)), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=helper_col + 8, min_row=status_start + 1, max_row=status_start + len(top_status)))
    ws.add_chart(bar, "G56")

    section(ws, 77, "5. Ganho operacional estimado", "Para traduzir tecnologia em impacto: abaixo usamos 5 minutos como referência de trabalho manual por nota. A premissa pode ser alterada na aba “Ganho Operacional”.")
    manual_minutes = 5
    system_minutes = (1 / notes_minute) if notes_minute else 0.1
    manual_hours = total_notes * manual_minutes / 60
    system_hours = total_notes * system_minutes / 60
    saved_hours = max(0, manual_hours - system_hours)
    reduction = saved_hours / manual_hours if manual_hours else 0
    card(ws, 1, 81, "VELOCIDADE MÉDIA", notes_minute, "Notas processadas por minuto nos históricos", BLUE, "0.00")
    card(ws, 4, 81, "HORAS MANUAIS EVITADAS", saved_hours, "Estimativa com base em 5 minutos por nota", GREEN, "#,##0")
    card(ws, 7, 81, "REDUÇÃO DE ESFORÇO", reduction, "Tempo operacional potencialmente reduzido", CYAN, "0.0%")
    card(ws, 10, 81, "PROCESSANDO AGORA", running_certs, "Certificados em execução no momento do extrato", AMBER)

    ws.merge_cells("A89:L92")
    ws["A89"] = "Como apresentar: comece pelos quatro resultados principais, mostre a evolução mensal, explique a divisão entre tomadas e prestadas e finalize com o ganho operacional. Os valores de economia são estimativas; os volumes, usuários, certificados e arquivos vêm diretamente da produção."
    ws["A89"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A89"].font = Font(size=11, color=WHITE)
    ws["A89"].alignment = Alignment(wrap_text=True, vertical="center")

    for col in range(helper_col, helper_col + 12):
        ws.column_dimensions[get_column_letter(col)].hidden = True
    ws.auto_filter.ref = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.tabColor = BLUE

    if "Dashboard Executivo" in wb.sheetnames:
        wb["Dashboard Executivo"].title = "Painel Detalhado"
    wb.active = 0
    wb.save(OUTPUT)
    print(f"Arquivo criado: {OUTPUT}")


if __name__ == "__main__":
    main()
