from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from sqlalchemy import func

from backend.app.db.models import AcessoUsuario, Arquivo, Certificado, Empresa, Evento, Nota, Processo, Usuario
from backend.app.db.session import SessionLocal
from backend.app.services.legacy_ingestion_service import _parse_xml_resumo


ROOT = Path(__file__).resolve().parents[3]
STORAGE = ROOT / "storage"
OUTPUT = ROOT / "relatorios" / "Apresentacao_Resultados_Portal_NFSe.xlsx"

NAVY = "12233F"
BLUE = "1D4ED8"
CYAN = "06B6D4"
GREEN = "10B981"
AMBER = "F59E0B"
RED = "EF4444"
LIGHT = "F3F6FA"
WHITE = "FFFFFF"
GRAY = "64748B"
GRID = "D7DFEA"


def number(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(Decimal(str(value).replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0.0


def month_key(value: str, fallback: Path) -> str:
    text = str(value or "")[:10]
    if len(text) >= 7 and text[4] == "-":
        return text[:7]
    parts = fallback.parts
    try:
        idx = parts.index("xml")
        return f"{int(parts[idx + 2]):04d}-{int(parts[idx + 3]):02d}"
    except (ValueError, IndexError):
        return "Sem data"


def company_cnpj(path: Path) -> str:
    try:
        idx = path.parts.index("xml")
        return path.parts[idx + 1]
    except (ValueError, IndexError):
        return "Não identificado"


def db_metrics() -> dict[str, int | float | None]:
    result: dict[str, int | float | None] = {
        "usuarios": None,
        "usuarios_ativos": None,
        "acessos": None,
        "certificados": None,
        "certificados_ativos": None,
        "certificados_rodando": None,
        "empresas": None,
        "processos": None,
        "notas_por_minuto": None,
    }
    try:
        with SessionLocal() as db:
            result.update(
                usuarios=db.query(Usuario).count(),
                usuarios_ativos=db.query(Usuario).filter(Usuario.ativo.is_(True)).count(),
                acessos=db.query(AcessoUsuario).count(),
                certificados=db.query(Certificado).count(),
                certificados_ativos=db.query(Certificado).filter(Certificado.ativo.is_(True)).count(),
                certificados_rodando=db.query(func.count(func.distinct(Processo.certificado_id))).filter(Processo.status == "rodando", Processo.certificado_id.isnot(None)).scalar(),
                empresas=db.query(Empresa).count(),
                processos=db.query(Processo).count(),
            )
            durations = db.query(Processo.started_at, Processo.finished_at).filter(Processo.started_at.isnot(None), Processo.finished_at.isnot(None)).all()
            seconds = sum(max(0, (end - start).total_seconds()) for start, end in durations)
            linked_notes = db.query(Nota).filter(Nota.processo_id.isnot(None)).count()
            result["notas_por_minuto"] = linked_notes / (seconds / 60) if seconds else 0
    except Exception:
        pass
    return result


def database_inventory():
    rows = []
    with SessionLocal() as db:
        records = db.query(
            Empresa.cnpj,
            Nota.data_emissao,
            Nota.competencia,
            Nota.numero_nfse,
            Nota.chave,
            Nota.prestador_cnpj,
            Nota.prestador_nome,
            Nota.tomador_nome,
            Nota.municipio,
            Nota.incidencia_iss,
            Nota.status_rotulo,
            Nota.status_documento,
            Nota.valor_servico,
            Nota.valor_liquido,
            Nota.valor_iss_retido,
            Nota.alertas_fiscais,
            Nota.xml_storage_key,
        ).join(Empresa, Empresa.id == Nota.empresa_id).yield_per(1000)
        pdf_count = db.query(Arquivo).filter(func.lower(Arquivo.tipo).like("%pdf%")).count()
        total_files = db.query(Arquivo).count()
        events = db.query(Evento).count()
        for record in records:
            cnpj = record.cnpj
            issue_date = record.data_emissao or record.competencia
            rows.append({
                "cnpj_empresa": cnpj,
                "mes": issue_date.strftime("%Y-%m") if issue_date else "Sem data",
                "data_emissao": issue_date,
                "numero": record.numero_nfse or "",
                "chave": record.chave,
                "direcao": "Prestado" if record.prestador_cnpj == cnpj else "Tomado",
                "prestador": record.prestador_nome or "",
                "tomador": record.tomador_nome or "",
                "municipio": record.municipio or record.incidencia_iss or "",
                "status": record.status_rotulo or record.status_documento or "Não informado",
                "valor_servico": number(record.valor_servico),
                "valor_liquido": number(record.valor_liquido),
                "iss_retido": number(record.valor_iss_retido),
                "alerta": "Sim" if record.alertas_fiscais else "Não",
                "arquivo": record.xml_storage_key or "",
            })
    return rows, len(rows), pdf_count, events, 0, total_files


def inventory():
    xml_files = sorted((STORAGE / "xml").rglob("*.xml"))
    pdf_files = list(STORAGE.rglob("*.pdf"))
    rows = []
    seen = set()
    errors = 0
    events = 0
    for path in xml_files:
        try:
            item = _parse_xml_resumo(path)
        except Exception:
            errors += 1
            continue
        if item.get("tipo_xml") == "evento":
            events += 1
            continue
        cnpj = company_cnpj(path)
        key = item.get("chave") or str(path.relative_to(STORAGE))
        unique = (cnpj, key)
        if unique in seen:
            continue
        seen.add(unique)
        direction = "Prestado" if item.get("prestador_cnpj") == cnpj else "Tomado"
        rows.append(
            {
                "cnpj_empresa": cnpj,
                "mes": month_key(item.get("data_emissao") or item.get("competencia"), path),
                "data_emissao": item.get("data_emissao") or item.get("competencia") or "",
                "numero": item.get("numero_nfse") or "",
                "chave": key,
                "direcao": direction,
                "prestador": item.get("prestador_nome") or "",
                "tomador": item.get("tomador_nome") or "",
                "municipio": item.get("municipio") or item.get("incidencia_iss") or "",
                "status": item.get("status_rotulo") or item.get("status_documento") or "Não informado",
                "valor_servico": number(item.get("valor_servico")),
                "valor_liquido": number(item.get("valor_liquido")),
                "iss_retido": number(item.get("valor_iss_retido")),
                "alerta": "Sim" if item.get("alertas_fiscais") else "Não",
                "arquivo": str(path.relative_to(ROOT)),
            }
        )
    return rows, len(xml_files), len(pdf_files), events, errors


def title(ws, text: str, subtitle: str | None = None):
    ws.sheet_view.showGridLines = False
    ws["A1"] = text
    ws["A1"].font = Font(size=24, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:L2")
    ws["A1"].alignment = Alignment(vertical="center")
    if subtitle:
        ws["A3"] = subtitle
        ws["A3"].font = Font(size=11, color=GRAY, italic=True)
        ws.merge_cells("A3:L3")


def kpi(ws, cell: str, label: str, value, color: str, fmt: str = "#,##0"):
    c = ws[cell]
    c.value = label
    c.font = Font(size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center")
    value_cell = ws.cell(c.row + 1, c.column)
    value_cell.value = value
    value_cell.font = Font(size=20, bold=True, color=NAVY)
    value_cell.fill = PatternFill("solid", fgColor=WHITE)
    value_cell.alignment = Alignment(horizontal="center")
    value_cell.number_format = fmt
    for row in range(c.row, c.row + 2):
        for col in range(c.column, c.column + 2):
            x = ws.cell(row, col)
            x.border = Border(left=Side(color=GRID), right=Side(color=GRID), top=Side(color=GRID), bottom=Side(color=GRID))
    ws.merge_cells(start_row=c.row, start_column=c.column, end_row=c.row, end_column=c.column + 1)
    ws.merge_cells(start_row=c.row + 1, start_column=c.column, end_row=c.row + 1, end_column=c.column + 1)


def table(ws, start_row: int, headers: list[str], data: list[list], name: str):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
    for row_idx, row in enumerate(data, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)
    if data:
        ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(data)}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
        ws.add_table(tab)
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(data)}"


def build(source: str = "local"):
    if source == "database":
        rows, raw_xml, pdf_count, events, errors, total_files = database_inventory()
        source_label = "banco de produção do Portal NFS-e"
        output = OUTPUT.with_name("Apresentacao_Resultados_Portal_NFSe_Online.xlsx")
    else:
        rows, raw_xml, pdf_count, events, errors = inventory()
        total_files = raw_xml + pdf_count
        source_label = "acervo local de XML/PDF"
        output = OUTPUT
    db = db_metrics()
    months = defaultdict(lambda: {"notas": 0, "valor": 0.0, "tomadas": 0, "prestadas": 0})
    companies = defaultdict(lambda: {"notas": 0, "valor": 0.0})
    statuses = Counter()
    for row in rows:
        months[row["mes"]]["notas"] += 1
        months[row["mes"]]["valor"] += row["valor_servico"]
        months[row["mes"]]["tomadas" if row["direcao"] == "Tomado" else "prestadas"] += 1
        companies[row["cnpj_empresa"]]["notas"] += 1
        companies[row["cnpj_empresa"]]["valor"] += row["valor_servico"]
        statuses[row["status"]] += 1

    total = len(rows)
    total_value = sum(r["valor_servico"] for r in rows)
    coverage = min(pdf_count / total, 1) if total else 0
    first_month = min((m for m in months if m != "Sem data"), default="-")
    last_month = max((m for m in months if m != "Sem data"), default="-")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Executivo"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    title(ws, "Portal NFS-e | Resultados e Ganhos", f"Fonte: {source_label} • período {first_month} a {last_month} • gerado em {datetime.now():%d/%m/%Y %H:%M}")
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14
    kpi(ws, "A5", "NFS-e únicas processadas", total, BLUE)
    kpi(ws, "C5", "Valor de serviços", total_value, GREEN, 'R$ #,##0.00')
    kpi(ws, "E5", "Empresas no acervo", len(companies), CYAN)
    kpi(ws, "G5", "PDFs disponíveis", pdf_count, AMBER)
    kpi(ws, "I5", "Cobertura documental", coverage, BLUE, "0.0%")
    kpi(ws, "K5", "Eventos processados", events, GREEN)
    kpi(ws, "A9", "Usuários cadastrados*", db["usuarios"], CYAN)
    kpi(ws, "C9", "Acessos registrados*", db["acessos"], BLUE)
    kpi(ws, "E9", "Certificados ativos*", db["certificados_ativos"], GREEN)
    kpi(ws, "G9", "Certificados rodando*", db["certificados_rodando"], AMBER)
    kpi(ws, "I9", "Arquivos registrados", total_files, BLUE)
    kpi(ws, "K9", "Notas por minuto", db["notas_por_minuto"], GREEN, "0.00")

    evo = wb.create_sheet("Evolução Mensal")
    month_rows = [[m, v["notas"], v["tomadas"], v["prestadas"], v["valor"]] for m, v in sorted(months.items())]
    title(evo, "Evolução mensal", "Volume de NFS-e únicas e valor de serviços por mês de emissão.")
    table(evo, 5, ["Mês", "Notas", "Tomadas", "Prestadas", "Valor dos serviços"], month_rows, "tbEvolucao")
    evo.column_dimensions["A"].width = 14
    evo.column_dimensions["E"].width = 22
    for cell in evo["E"][5:]:
        cell.number_format = 'R$ #,##0.00'

    line = LineChart()
    line.title = "Notas processadas por mês"
    line.y_axis.title = "Quantidade"
    line.x_axis.title = "Mês"
    line.style = 13
    line.height, line.width = 8, 17
    line.add_data(Reference(evo, min_col=2, max_col=4, min_row=5, max_row=5 + len(month_rows)), titles_from_data=True)
    line.set_categories(Reference(evo, min_col=1, min_row=6, max_row=5 + len(month_rows)))
    ws.add_chart(line, "A13")

    status_ws = wb.create_sheet("Status e Qualidade")
    status_rows = [[k, v, v / total if total else 0] for k, v in statuses.most_common()]
    title(status_ws, "Status e qualidade documental", "Distribuição dos documentos conforme o status identificado no XML.")
    table(status_ws, 5, ["Status", "Quantidade", "% do total"], status_rows, "tbStatus")
    status_ws.column_dimensions["A"].width = 28
    status_ws.column_dimensions["B"].width = 16
    status_ws.column_dimensions["C"].width = 16
    for cell in status_ws["C"][5:]:
        cell.number_format = "0.0%"
    pie = DoughnutChart()
    pie.title = "Distribuição por status"
    pie.height, pie.width = 8, 13
    pie.add_data(Reference(status_ws, min_col=2, min_row=5, max_row=5 + len(status_rows)), titles_from_data=True)
    pie.set_categories(Reference(status_ws, min_col=1, min_row=6, max_row=5 + len(status_rows)))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "G13")

    emp = wb.create_sheet("Empresas")
    company_rows = [[k, v["notas"], v["valor"], v["notas"] / total if total else 0] for k, v in sorted(companies.items(), key=lambda x: x[1]["notas"], reverse=True)]
    title(emp, "Resultados por empresa", "Participação de cada CNPJ no acervo processado.")
    table(emp, 5, ["CNPJ da empresa", "Notas", "Valor dos serviços", "% do volume"], company_rows, "tbEmpresas")
    for col, width in {"A": 22, "B": 14, "C": 22, "D": 16}.items():
        emp.column_dimensions[col].width = width
    for cell in emp["C"][5:]: cell.number_format = 'R$ #,##0.00'
    for cell in emp["D"][5:]: cell.number_format = "0.0%"

    gain = wb.create_sheet("Ganho Operacional")
    title(gain, "Ganho operacional estimado", "Premissas editáveis para simular o esforço manual evitado pela automação.")
    gain.column_dimensions["A"].width = 42
    gain.column_dimensions["B"].width = 22
    gain.column_dimensions["C"].width = 55
    inputs = [
        ("Notas processadas", total, "Calculado a partir das NFS-e únicas do acervo"),
        ("Minutos por nota no processo manual", 5, "Premissa editável: localizar, baixar, organizar e conferir"),
        ("Minutos por nota com o sistema", (1 / float(db["notas_por_minuto"])) if db["notas_por_minuto"] else 0.1, "Calculado pela duração registrada dos processos; pode ser ajustado"),
        ("Custo médio por hora (R$)", 35, "Premissa editável para estimar economia financeira"),
        ("Horas no processo manual", "=B5*B6/60", "Estimativa"),
        ("Horas com o sistema", "=B5*B7/60", "Estimativa"),
        ("Horas economizadas", "=B9-B10", "Ganho estimado"),
        ("Redução de esforço", "=IFERROR(B11/B9,0)", "Percentual estimado"),
        ("Economia financeira", "=B11*B8", "Estimativa"),
        ("Capacidade manual (notas/hora)", "=60/B6", "Capacidade estimada"),
        ("Capacidade com sistema (notas/hora)", "=60/B7", "Capacidade estimada"),
        ("Multiplicação de capacidade", "=B15/B14", "Ganho estimado"),
    ]
    table(gain, 4, ["Indicador / premissa", "Valor", "Observação"], [list(x) for x in inputs], "tbGanho")
    for row in (6, 7, 8):
        gain[f"B{row}"].fill = PatternFill("solid", fgColor="FFF3CD")
    gain["B12"].number_format = "0.0%"
    gain["B13"].number_format = 'R$ #,##0.00'
    gain["B16"].number_format = '0.0x'
    gain["E4"], gain["F4"] = "Cenário", "Horas"
    gain["E5"], gain["F5"] = "Manual", "=B9"
    gain["E6"], gain["F6"] = "Com sistema", "=B10"
    bar = BarChart()
    bar.type = "col"
    bar.title = "Esforço estimado: manual x sistema"
    bar.y_axis.title = "Horas"
    bar.height, bar.width = 8, 14
    bar.add_data(Reference(gain, min_col=6, min_row=4, max_row=6), titles_from_data=True)
    bar.set_categories(Reference(gain, min_col=5, min_row=5, max_row=6))
    gain.add_chart(bar, "E8")

    base = wb.create_sheet("Base de Notas")
    headers = ["CNPJ empresa", "Mês", "Data emissão", "Número", "Chave", "Direção", "Prestador", "Tomador", "Município", "Status", "Valor serviço", "Valor líquido", "ISS retido", "Alerta fiscal", "Arquivo de origem"]
    data = [[r[k] for k in ("cnpj_empresa", "mes", "data_emissao", "numero", "chave", "direcao", "prestador", "tomador", "municipio", "status", "valor_servico", "valor_liquido", "iss_retido", "alerta", "arquivo")] for r in rows]
    table(base, 1, headers, data, "tbNotas")
    widths = [20, 12, 14, 16, 52, 12, 34, 34, 25, 20, 18, 18, 16, 14, 70]
    for idx, width in enumerate(widths, 1): base.column_dimensions[get_column_letter(idx)].width = width
    for col in (11, 12, 13):
        for cell in base[get_column_letter(col)][1:]: cell.number_format = 'R$ #,##0.00'

    notes = wb.create_sheet("Fontes e Premissas")
    title(notes, "Fontes, cobertura e premissas", "Informações essenciais para apresentar os números com transparência.")
    notes.column_dimensions["A"].width = 34
    notes.column_dimensions["B"].width = 95
    notes_data = [
        ["Fonte de notas", source_label if source == "database" else str(STORAGE / "xml")],
        ["Critério de deduplicação", "Restrição única empresa + chave no banco" if source == "database" else "CNPJ da empresa + chave da NFS-e"],
        ["XMLs brutos", raw_xml],
        ["NFS-e únicas", total],
        ["Eventos XML separados", events],
        ["PDFs encontrados", pdf_count],
        ["Período", f"{first_month} a {last_month}"],
        ["Métricas com *", "Obtidas diretamente das tabelas de produção no momento da geração." if source == "database" else "Obtidas da base local atualmente configurada; valide/substitua pelos números do banco online."],
        ["Ganho operacional", "Estimativa baseada nas premissas amarelas da aba Ganho Operacional; ajuste-as à rotina real da equipe."],
        ["Velocidade", "A aba de ganho apresenta capacidade estimada. A velocidade real por execução exige histórico started_at/finished_at do banco online."],
        ["Atualização dos gráficos", "As tabelas e gráficos acompanham os dados consolidados desta geração. Para novo extrato, execute novamente o gerador."],
    ]
    table(notes, 5, ["Item", "Detalhe"], notes_data, "tbFontes")
    for row in notes.iter_rows(min_row=6, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    for sheet in wb.worksheets:
        sheet.sheet_view.zoomScale = 85
        sheet.freeze_panes = sheet.freeze_panes or "A4"
        sheet.auto_filter.ref = sheet.auto_filter.ref
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Arquivo criado: {output}")
    print(f"NFS-e únicas: {total}; XMLs: {raw_xml}; PDFs: {pdf_count}; empresas: {len(companies)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", choices=("local", "database"), default="local")
    args = parser.parse_args()
    build(args.source)
