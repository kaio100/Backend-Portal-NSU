from __future__ import annotations

import re
import unicodedata
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REGRAS_XLSX_PATH = Path(__file__).resolve().parents[3] / "data" / "RETENCOES_REGRAS.xlsx"


def _sem_acentos(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def normalizar_flag_regra(value: str | None) -> str:
    text = _sem_acentos(str(value or "")).strip().upper()
    if not text:
        return "NAO"
    if text in {"SIM", "S"}:
        return "SIM"
    if text in {"NAO", "N", "NO", "NÃO"}:
        return "NAO"
    if text == "DEPENDE":
        return "DEPENDE"
    return "DEPENDE" if "DEPENDE" in text else "NAO"


def normalizar_subitem_lc116(value: str | None) -> str | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 6:
        grupo = int(digits[:2])
        return f"{grupo}.{digits[2:4]}"
    if len(digits) == 4:
        grupo = int(digits[:2])
        return f"{grupo}.{digits[2:4]}"
    if len(digits) == 3:
        grupo = int(digits[:1])
        return f"{grupo}.{digits[1:3]}"
    match = re.search(r"(\d{1,2})\D+(\d{1,2})", raw)
    if match:
        return f"{int(match.group(1))}.{int(match.group(2)):02d}"
    return None


def _decimal_or_none(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).replace(",", "."))
    except InvalidOperation:
        return None


def _regra_from_row(row: dict[str, Any]) -> dict[str, Any] | None:
    subitem = normalizar_subitem_lc116(row.get("Subitem"))
    if not subitem:
        return None
    observacao = str(row.get("Observação fiscal essencial") or "").strip()
    if observacao in {"", "-", "—"}:
        observacao = ""
    obs_norm = _sem_acentos(observacao).lower()
    return {
        "subitem": subitem,
        "descricao": str(row.get("Descrição legal resumida") or "").strip(),
        "irrf": normalizar_flag_regra(row.get("IRRF")),
        "irrf_aliquota": _decimal_or_none(row.get("Alíquota")),
        "pcc": normalizar_flag_regra(row.get("PIS/COFINS/CSLL")),
        "inss": normalizar_flag_regra(row.get("INSS")),
        "observacao": observacao or None,
        "inss_optante_simples_retem": "optante do simples tambem retem" in obs_norm,
    }


@lru_cache(maxsize=1)
def carregar_regras_retencao() -> dict[str, dict[str, Any]]:
    if not REGRAS_XLSX_PATH.exists():
        return {}
    wb = load_workbook(REGRAS_XLSX_PATH, data_only=True, read_only=True)
    ws = wb["Planilha2"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    regras: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        regra = _regra_from_row(dict(zip(headers, row)))
        if regra:
            regras[regra["subitem"]] = regra
    return regras


def obter_regra_por_subitem(subitem: str | None) -> dict[str, Any] | None:
    normalized = normalizar_subitem_lc116(subitem)
    if not normalized:
        return None
    return carregar_regras_retencao().get(normalized)


def resolver_subitem_lc116(dados_xml: dict[str, Any]) -> str | None:
    for key in ("cTribNac", "codigo_servico_nacional", "codigo_servico", "cServ", "cServMun", "ItemListaServico", "itemListaServico", "CodigoServico", "CodigoTributacaoMunicipio"):
        subitem = normalizar_subitem_lc116(dados_xml.get(key))
        if subitem:
            return subitem
    return None
