from __future__ import annotations

import io
import os
import re
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import or_
from sqlalchemy.orm import Session

from backend.app.core.config import settings
from backend.app.db.models import Arquivo, Empresa, Evento, Nota, NsuControle, Processo
from backend.app.repositories import arquivos_repo, notas_repo, processos_repo
from backend.app.schemas.notas import NotasDownloadFiltros
from backend.app.services import cnpj_cache_service, legacy_ingestion_service, notas_service
from backend.app.services.operational_fields_service import aplicar_campos_operacionais
from backend.app.services.storage_service import StorageService, get_storage_service


class PortalSupportError(RuntimeError):
    pass


def _canonical_tipo_arquivo(tipo: str | None) -> str:
    normalized = (tipo or "").lower()
    if normalized == "xml":
        return "XML"
    if normalized in {"pdf_original", "pdf_oficial", "oficial"}:
        return "PDF_ORIGINAL"
    if normalized in {"pdf_espelho", "espelho"}:
        return "PDF_ESPELHO"
    return tipo or "OUTRO"


def _tipo_documento_frontend(tipo: str | None) -> str:
    canonical = _canonical_tipo_arquivo(tipo)
    if canonical == "XML":
        return "xml"
    if canonical in {"PDF_ORIGINAL", "PDF_ESPELHO"}:
        return "pdf"
    return canonical.lower()


def _matches_tipo_arquivo(arquivo: Arquivo, tipo: str) -> bool:
    tipo_norm = tipo.lower()
    canonical = _canonical_tipo_arquivo(arquivo.tipo).lower()
    frontend = _tipo_documento_frontend(arquivo.tipo)
    filename = (_nome_arquivo(arquivo) or "").lower()
    content_type = (arquivo.content_type or "").lower()
    if tipo_norm in {canonical, frontend}:
        return True
    if tipo_norm == "xlsx":
        return filename.endswith(".xlsx") or "spreadsheet" in content_type
    if tipo_norm == "csv":
        return filename.endswith(".csv") or "csv" in content_type
    if tipo_norm == "zip":
        return filename.endswith(".zip") or "zip" in content_type
    if tipo_norm == "relatorio":
        return canonical == "relatorio" or filename.endswith((".xlsx", ".csv"))
    if tipo_norm == "outro":
        return frontend not in {"xml", "pdf"} and canonical not in {"relatorio"}
    return False


def _nome_arquivo(arquivo: Arquivo) -> str:
    if arquivo.filename:
        return arquivo.filename
    return (arquivo.storage_key or "").replace("\\", "/").split("/")[-1] or f"arquivo_{arquivo.id}"


def _arquivo_disponivel(storage: StorageService | None, arquivo: Arquivo) -> bool:
    if storage is None:
        return True
    try:
        return storage.exists(arquivo.storage_key)
    except Exception:
        return False


def arquivo_item(arquivo: Arquivo, storage: StorageService | None = None) -> dict:
    return {
        "id": int(arquivo.id),
        "arquivo_id": int(arquivo.id),
        "nota_id": arquivo.nota_id,
        "processo_id": arquivo.processo_id,
        "tipo": _tipo_documento_frontend(arquivo.tipo),
        "tipo_canonico": _canonical_tipo_arquivo(arquivo.tipo),
        "nome": _nome_arquivo(arquivo),
        "filename": _nome_arquivo(arquivo),
        "content_type": arquivo.content_type,
        "tamanho": arquivo.tamanho_bytes,
        "size_bytes": arquivo.tamanho_bytes,
        "disponivel": _arquivo_disponivel(storage, arquivo),
        "download_url": None,
        "criado_em": arquivo.created_at,
        "created_at": arquivo.created_at,
    }


def listar_documentos_nota(db: Session, nota_id: int, storage: StorageService | None = None) -> dict:
    notas_service.obter_nota(db, nota_id)
    arquivos = notas_service.listar_arquivos_nota(db, nota_id)
    return {
        "nota_id": nota_id,
        "items": [arquivo_item(arquivo, storage) for arquivo in arquivos],
        "total": len(arquivos),
    }


def _arquivo_id_por_storage(db: Session, storage_key: str | None) -> int | None:
    if not storage_key:
        return None
    arquivo = arquivos_repo.get_arquivo_by_storage_key(db, storage_key)
    return int(arquivo.id) if arquivo is not None else None


def evento_item(db: Session, evento: Evento) -> dict:
    return {
        "id": int(evento.id),
        "tipo_evento": evento.tipo_evento,
        "codigo_evento": evento.tipo_evento,
        "descricao": evento.descricao,
        "data_evento": evento.data_evento or evento.created_at,
        "protocolo": None,
        "chave_afetada": evento.chave_afetada,
        "status": evento.tipo_evento,
        "arquivo_id": _arquivo_id_por_storage(db, evento.xml_storage_key),
        "nsu": evento.nsu,
    }


def listar_eventos_nota(db: Session, nota_id: int) -> dict:
    nota = notas_service.obter_nota(db, nota_id)
    eventos = (
        db.query(Evento)
        .filter(or_(Evento.nota_id == nota_id, Evento.chave_afetada == nota.chave))
        .order_by(Evento.data_evento.desc().nullslast(), Evento.created_at.desc(), Evento.id.desc())
        .all()
    )
    return {
        "nota_id": nota_id,
        "items": [evento_item(db, evento) for evento in eventos],
        "total": len(eventos),
    }


def listar_eventos(
    db: Session,
    empresa_id: int | None = None,
    nota_id: int | None = None,
    chave_afetada: str | None = None,
    tipo_evento: str | None = None,
    data_inicio: date | None = None,
    data_fim: date | None = None,
    limit: int = 100,
    offset: int = 0,
) -> dict:
    query = db.query(Evento)
    if empresa_id is not None:
        query = query.filter(Evento.empresa_id == empresa_id)
    if nota_id is not None:
        query = query.filter(Evento.nota_id == nota_id)
    if chave_afetada:
        query = query.filter(Evento.chave_afetada == chave_afetada)
    if tipo_evento:
        query = query.filter(Evento.tipo_evento == tipo_evento)
    if data_inicio is not None:
        query = query.filter(Evento.data_evento >= data_inicio)
    if data_fim is not None:
        query = query.filter(Evento.data_evento <= data_fim)
    total = query.count()
    eventos = (
        query.order_by(Evento.data_evento.desc().nullslast(), Evento.created_at.desc(), Evento.id.desc())
        .offset(max(offset, 0))
        .limit(min(max(limit, 1), 500))
        .all()
    )
    return {"items": [evento_item(db, evento) for evento in eventos], "total": total}


# Status que representam "sem problema" - nao ha nada a corrigir, entao a
# coluna de observacao do comparativo de tributos deve ficar vazia. Qualquer
# outro status (Divergente, Divergente - retido indevido, Divergente - nao
# retido, Depende de analise, etc.) mostra a observacao de alerta.
_STATUS_TRIBUTO_SEM_PROBLEMA = {
    "ok",
    "correto",
    "nao se aplica",
    "nao retido",
    "nao informado",
    "informado",
}


def comparar_tributos_nota(db: Session, nota_id: int) -> dict:
    nota = notas_service.obter_nota(db, nota_id)
    specs = [
        ("IRRF", "irrf", "irrf_calculado", "status_irrf", "IRRF esperado diferente do informado"),
        ("CSRF", "csrf", "csrf_calculado", "status_csrf", "CSRF esperado diferente do informado"),
        ("INSS", "inss", "inss_calculado", "status_inss", "INSS esperado diferente do informado"),
        ("ISS", "iss", "iss_calculado", "status_iss", "ISS esperado diferente do informado"),
        ("VALOR_LIQUIDO", "valor_liquido", "valor_liquido_correto", "status_valor_liquido", "Valor liquido esperado diferente do informado"),
    ]
    items: list[dict] = []
    for tributo, informado_field, calculado_field, status_field, observacao in specs:
        informado = getattr(nota, informado_field, None)
        calculado = getattr(nota, calculado_field, None) if calculado_field else None
        status = getattr(nota, status_field, None) if status_field else None
        if informado is None and calculado is None and not status:
            continue
        informado_dec = _decimal(informado) if informado is not None else Decimal("0")
        calculado_dec = _decimal(calculado) if calculado is not None else informado_dec
        diferenca = calculado_dec - informado_dec
        final_status = status or ("Correto" if abs(diferenca) < Decimal("0.01") else "Divergente")
        sem_problema = final_status.strip().lower() in _STATUS_TRIBUTO_SEM_PROBLEMA
        items.append(
            {
                "tributo": tributo,
                "informado": float(informado_dec),
                "calculado": float(calculado_dec),
                "diferenca": float(diferenca),
                "status": final_status,
                "observacao": None if sem_problema else observacao,
            }
        )
    return {"nota_id": nota_id, "items": items}


def listar_arquivos_processo(
    db: Session,
    processo_id: int,
    tipo: str | None = None,
    storage: StorageService | None = None,
) -> dict:
    processo = processos_repo.get_processo(db, processo_id)
    if processo is None:
        raise PortalSupportError("Processo nao encontrado.")
    arquivos = arquivos_repo.list_arquivos(db, processo_id=processo_id, limit=500, offset=0)
    if tipo:
        arquivos = [arquivo for arquivo in arquivos if _matches_tipo_arquivo(arquivo, tipo)]
    items = [arquivo_item(arquivo, storage) for arquivo in arquivos]
    return {"processo_id": processo_id, "items": items, "total": len(items)}


def listar_notas_processo(
    db: Session,
    processo_id: int,
    status: str | None = None,
    conferencia_status: str | None = None,
    tipo_nota: str | None = None,
    direcao_nota: str | None = None,
    busca: str | None = None,
    somente_divergentes: bool = False,
    valor_min: Decimal | None = None,
    valor_max: Decimal | None = None,
    limit: int = 100,
    offset: int = 0,
) -> dict:
    processo = processos_repo.get_processo(db, processo_id)
    if processo is None:
        raise PortalSupportError("Processo nao encontrado.")
    notas = notas_service.listar_notas(
        db,
        processo_id=processo_id,
        status_documento=status,
        conferencia_status=conferencia_status,
        tipo_nota=tipo_nota,
        direcao_nota=direcao_nota,
        busca=busca,
        limit=5000,
        offset=0,
    )
    notas = _filtrar_notas_operacionais(
        notas,
        empresa_cnpj=(db.get(Empresa, processo.empresa_id).cnpj if db.get(Empresa, processo.empresa_id) else ""),
        somente_divergentes=somente_divergentes,
        valor_min=valor_min,
        valor_max=valor_max,
    )
    total = len(notas)
    safe_offset = max(offset, 0)
    safe_limit = min(max(limit, 1), 500)
    return {
        "processo_id": processo_id,
        "items": notas[safe_offset : safe_offset + safe_limit],
        "total": total,
        "limit": safe_limit,
        "offset": safe_offset,
    }


def _tem_divergencia(nota: Nota) -> bool:
    if nota.divergencia:
        return True
    if (nota.conferencia_status or "").lower() == "corrigir":
        return True
    for field in ("status_valor_liquido", "status_csrf", "status_irrf", "status_inss"):
        value = (getattr(nota, field, None) or "").lower()
        if value and value not in {"ok", "sem_divergencia", "regular"}:
            return True
    return False


def _filtrar_notas_operacionais(
    notas: list[Nota],
    empresa_cnpj: str = "",
    tipo_nota: str | None = None,
    somente_divergentes: bool = False,
    valor_min: Decimal | None = None,
    valor_max: Decimal | None = None,
) -> list[Nota]:
    filtradas = list(notas)
    if tipo_nota in {"emitida", "recebida"} and empresa_cnpj:
        filtradas = [
            nota for nota in filtradas
            if getattr(notas_service, "_classificar_nota")(nota, empresa_cnpj) == tipo_nota
        ]
    if somente_divergentes:
        filtradas = [nota for nota in filtradas if _tem_divergencia(nota)]
    if valor_min is not None:
        filtradas = [nota for nota in filtradas if _decimal(nota.valor_servico) >= Decimal(str(valor_min))]
    if valor_max is not None:
        filtradas = [nota for nota in filtradas if _decimal(nota.valor_servico) <= Decimal(str(valor_max))]
    return filtradas


def _decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


def resumo_processo(db: Session, processo_id: int) -> dict:
    processo = processos_repo.get_processo(db, processo_id)
    if processo is None:
        raise PortalSupportError("Processo nao encontrado.")
    notas = [aplicar_campos_operacionais(nota) for nota in db.query(Nota).filter(Nota.processo_id == processo_id).all()]
    arquivos = arquivos_repo.list_arquivos(db, processo_id=processo_id, limit=5000, offset=0)
    status_values = [(nota.status_documento or "").lower() for nota in notas]
    return {
        "processo_id": processo_id,
        "total_notas": len(notas),
        "total_xml": sum(1 for arquivo in arquivos if _canonical_tipo_arquivo(arquivo.tipo) == "XML"),
        "total_pdf": sum(1 for arquivo in arquivos if _canonical_tipo_arquivo(arquivo.tipo) in {"PDF_ORIGINAL", "PDF_ESPELHO"}),
        "corretas": sum(1 for nota in notas if getattr(nota, "status_fila_final", None) == "correta"),
        "divergentes": sum(1 for nota in notas if getattr(nota, "divergencia_fila_final", False)),
        "pendentes": sum(1 for nota in notas if getattr(nota, "status_fila_final", None) == "pendente"),
        "canceladas": sum(1 for status in status_values if "cancel" in status),
        "substituidas": sum(1 for status in status_values if "substit" in status),
        "valor_total_servicos": float(sum((_decimal(nota.valor_servico) for nota in notas), Decimal("0"))),
        "valor_total_iss": 0.0,
        "nsu_inicio": processo.nsu_inicio,
        "nsu_final": processo.nsu_final,
    }


def resumo_operacional_empresas(
    db: Session,
    data_inicio: date | None = None,
    data_fim: date | None = None,
    competencia_inicio: date | None = None,
    competencia_fim: date | None = None,
    status: str | None = None,
    conferencia_status: str | None = None,
) -> dict:
    items: list[dict] = []
    empresas = db.query(Empresa).order_by(Empresa.nome.asc()).all()
    for empresa in empresas:
        query = db.query(Nota).filter(Nota.empresa_id == empresa.id)
        if data_inicio is not None:
            query = query.filter(Nota.data_emissao >= data_inicio)
        if data_fim is not None:
            query = query.filter(Nota.data_emissao <= data_fim)
        if competencia_inicio is not None:
            query = query.filter(Nota.competencia >= competencia_inicio)
        if competencia_fim is not None:
            query = query.filter(Nota.competencia <= competencia_fim)
        if status:
            query = query.filter(Nota.status_documento == status)
        if conferencia_status:
            query = query.filter(Nota.conferencia_status == conferencia_status)
        notas = [aplicar_campos_operacionais(nota) for nota in query.all()]
        ultimo_processo = (
            db.query(Processo)
            .filter(Processo.empresa_id == empresa.id)
            .order_by(Processo.updated_at.desc().nullslast(), Processo.id.desc())
            .first()
        )
        ultimo_nsu = (
            db.query(NsuControle)
            .filter(NsuControle.empresa_id == empresa.id)
            .order_by(NsuControle.ultimo_nsu.desc())
            .first()
        )
        items.append(
            {
                "empresa_id": int(empresa.id),
                "empresa_nome": empresa.nome,
                "cnpj": empresa.cnpj,
                "total_notas": len(notas),
                "corretas": sum(1 for nota in notas if getattr(nota, "status_fila_final", None) == "correta"),
                "divergentes": sum(1 for nota in notas if getattr(nota, "divergencia_fila_final", False)),
                "pendentes": sum(1 for nota in notas if getattr(nota, "status_fila_final", None) == "pendente"),
                "ultima_execucao": ultimo_processo.updated_at if ultimo_processo is not None else None,
                "ultimo_status": ultimo_processo.status if ultimo_processo is not None else None,
                "ultimo_nsu": ultimo_nsu.ultimo_nsu if ultimo_nsu is not None else None,
            }
        )
    return {"items": items, "total": len(items)}


def _format_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _xlsx_value(value):
    if value is None:
        return 0
    if isinstance(value, str) and value.strip() == "":
        return 0
    if isinstance(value, Decimal):
        return float(value)
    return value


def _buscar_notas_relatorio(db: Session, filtros: NotasDownloadFiltros) -> list[Nota]:
    notas = notas_service.listar_notas(
        db,
        empresa_id=filtros.empresa_id,
        certificado_id=filtros.certificado_id,
        processo_id=filtros.processo_id,
        status_documento=filtros.status_documento,
        numero=filtros.numero,
        prestador_cnpj=filtros.prestador_cnpj,
        tomador_cnpj=filtros.tomador_cnpj,
        chave=filtros.chave,
        busca=filtros.busca,
        data_inicio=filtros.data_inicio,
        data_fim=filtros.data_fim,
        competencia_inicio=filtros.competencia_inicio,
        competencia_fim=filtros.competencia_fim,
        conferencia_status=filtros.conferencia_status,
        prioridade=filtros.prioridade,
        responsavel=filtros.responsavel,
        status_nota_pdf=filtros.status_nota_pdf,
        simples_nacional_xml=filtros.simples_nacional_xml,
        consulta_simples_api=None,
        status_simples_nacional=filtros.status_simples_nacional,
        incidencia_iss=filtros.incidencia_iss,
        divergencia=filtros.divergencia,
        sla_status=filtros.sla_status,
        tipo_nota=filtros.tipo_nota,
        direcao_nota=filtros.direcao_nota,
        limit=5000,
        offset=0,
        sort=filtros.sort,
    )
    empresa_cnpj = ""
    if filtros.empresa_id:
        empresa = db.get(Empresa, filtros.empresa_id)
        empresa_cnpj = empresa.cnpj if empresa is not None else ""
    return _filtrar_notas_operacionais(
        notas,
        empresa_cnpj=empresa_cnpj,
        tipo_nota=filtros.tipo_nota,
        somente_divergentes=filtros.somente_divergentes,
        valor_min=filtros.valor_min,
        valor_max=filtros.valor_max,
    )


def _parse_decimal_optional(value) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).replace(",", "."))
    except InvalidOperation:
        return None


def _is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def _nota_field(nota: Nota, field: str, xml_resumo: dict[str, str] | None = None, xml_key: str | None = None):
    value = getattr(nota, field, None)
    if not _is_blank(value):
        return value
    if xml_resumo is None:
        return value
    return xml_resumo.get(xml_key or field) or value


def _nota_decimal_field(nota: Nota, field: str, xml_resumo: dict[str, str] | None = None, xml_key: str | None = None):
    value = getattr(nota, field, None)
    if not _is_blank(value):
        return value
    if xml_resumo is None:
        return value
    return _parse_decimal_optional(xml_resumo.get(xml_key or field))


def _status_valor_liquido(nota: Nota, xml_resumo: dict[str, str] | None = None) -> str:
    status = getattr(nota, "status_valor_liquido", None)
    if not _is_blank(status):
        return _format_value(status)
    if xml_resumo and not _is_blank(xml_resumo.get("status_valor_liquido")):
        return _format_value(xml_resumo.get("status_valor_liquido"))
    if not xml_resumo or not xml_resumo.get("valor_liquido"):
        return ""
    valor_relatorio = _parse_decimal_optional(getattr(nota, "valor_liquido", None))
    valor_xml = _parse_decimal_optional(xml_resumo.get("valor_liquido"))
    if valor_relatorio is None or valor_xml is None:
        return ""
    return "OK" if valor_relatorio.quantize(Decimal("0.01")) == valor_xml.quantize(Decimal("0.01")) else "Divergente"


def _nota_relatorio_row(nota: Nota, xml_resumo: dict[str, str] | None = None) -> list:
    sla = getattr(nota, "sla", None)
    sla_label = sla.get("label") if isinstance(sla, dict) else _format_value(sla)
    return [
        _format_value(nota.competencia),
        _format_value(_nota_field(nota, "municipio", xml_resumo)),
        _format_value(nota.chave),
        _format_value(nota.data_emissao),
        _format_value(nota.prestador_cnpj),
        _format_value(nota.prestador_nome),
        _format_value(nota.numero_nfse),
        nota.valor_servico,
        _nota_decimal_field(nota, "valor_base", xml_resumo),
        _nota_decimal_field(nota, "csrf", xml_resumo),
        _nota_decimal_field(nota, "irrf", xml_resumo),
        _nota_decimal_field(nota, "inss", xml_resumo),
        _nota_decimal_field(nota, "iss", xml_resumo),
        _nota_decimal_field(nota, "aliquota_iss", xml_resumo),
        nota.valor_liquido,
        _nota_decimal_field(nota, "valor_liquido_correto", xml_resumo),
        _status_valor_liquido(nota, xml_resumo),
        ", ".join(getattr(nota, "campos_ausentes_xml", []) or []),
        _format_value(_nota_field(nota, "incidencia_iss", xml_resumo)),
        _format_value(_nota_field(nota, "codigo_servico", xml_resumo)),
        _format_value(_nota_field(nota, "descricao_servico_nacional", xml_resumo)),
        _format_value(_nota_field(nota, "descricao_servico_detalhada", xml_resumo)),
        _format_value(getattr(nota, "cnae", None)),
        _format_value(getattr(nota, "simples_nacional", None) or getattr(nota, "simples_xml", None) or nota.simples_nacional_xml),
        _format_value(nota.status_simples_nacional),
        _format_value(nota.status_documento),
        _format_value(getattr(nota, "divergencia_fila_label", None)),
        _format_value(getattr(nota, "prioridade_fila", None) or nota.prioridade or getattr(nota, "prioridade_manual", None)),
        _format_value(nota.responsavel),
        _format_value(sla_label),
        _format_value(nota.conferencia_observacao),
        _format_value(getattr(nota, "status_csrf", None)),
        _format_value(getattr(nota, "status_irrf", None)),
        _format_value(getattr(nota, "status_inss", None)),
        _format_value(_nota_field(nota, "subitem_lc116", xml_resumo)),
        _format_value(_nota_field(nota, "codigo_servico_nacional", xml_resumo)),
        _format_value(_nota_field(nota, "origem_base_calculo", xml_resumo)),
        _format_value(_nota_field(nota, "iss_retido", xml_resumo)),
        _nota_decimal_field(nota, "valor_iss_retido", xml_resumo),
        _nota_decimal_field(nota, "valor_pis", xml_resumo),
        _nota_decimal_field(nota, "pis_calculado", xml_resumo),
        _nota_decimal_field(nota, "valor_cofins", xml_resumo),
        _nota_decimal_field(nota, "cofins_calculado", xml_resumo),
        _nota_decimal_field(nota, "valor_csll", xml_resumo),
        _nota_decimal_field(nota, "csll_calculado", xml_resumo),
        _nota_decimal_field(nota, "csrf_calculado", xml_resumo),
        _nota_decimal_field(nota, "irrf_calculado", xml_resumo),
        _nota_decimal_field(nota, "inss_calculado", xml_resumo),
        _nota_decimal_field(nota, "iss_calculado", xml_resumo),
        _format_value(_nota_field(nota, "status_iss", xml_resumo)),
        _nota_decimal_field(nota, "valor_liquido_calculado", xml_resumo),
        _format_value(_nota_field(nota, "regra_irrf", xml_resumo)),
        _nota_decimal_field(nota, "regra_irrf_aliquota", xml_resumo),
        _format_value(_nota_field(nota, "regra_pcc", xml_resumo)),
        _format_value(_nota_field(nota, "regra_inss", xml_resumo)),
        _format_value(_nota_field(nota, "regra_observacao", xml_resumo)),
        _format_value(getattr(nota, "alertas_fiscais", None)),
    ]


def _precisa_enriquecimento_xml(nota: Nota) -> bool:
    fields = [
        "municipio",
        "valor_base",
        "csrf",
        "irrf",
        "inss",
        "iss",
        "aliquota_iss",
        "valor_liquido_correto",
        "status_valor_liquido",
        "incidencia_iss",
        "codigo_servico",
        "descricao_servico_nacional",
        "descricao_servico_detalhada",
    ]
    return any(_is_blank(getattr(nota, field, None)) for field in fields)


def _xml_resumos_por_nota(db: Session, notas: list[Nota]) -> dict[int, dict[str, str]]:
    notas_pendentes = [nota for nota in notas if _precisa_enriquecimento_xml(nota)]
    if not notas_pendentes:
        return {}
    try:
        storage = get_storage_service()
    except Exception:
        return {}

    arquivos_por_nota: dict[int, Arquivo] = {}
    arquivos = arquivos_repo.list_arquivos_by_notas(db, [int(nota.id) for nota in notas_pendentes])
    for arquivo in arquivos:
        if _canonical_tipo_arquivo(arquivo.tipo) == "XML" and arquivo.nota_id is not None:
            arquivos_por_nota.setdefault(int(arquivo.nota_id), arquivo)

    tarefas: list[tuple[int, str, str]] = []
    for nota in notas_pendentes:
        arquivo = arquivos_por_nota.get(int(nota.id))
        storage_key = arquivo.storage_key if arquivo is not None else getattr(nota, "xml_storage_key", None)
        if storage_key:
            filename = _nome_arquivo(arquivo) if arquivo is not None else storage_key.rsplit("/", 1)[-1]
            tarefas.append((int(nota.id), storage_key, filename))

    def carregar_xml(tarefa: tuple[int, str, str]):
        nota_id, storage_key, filename = tarefa
        try:
            resumo = legacy_ingestion_service.parse_xml_resumo_bytes(storage.get_bytes(storage_key), filename)
            return nota_id, resumo
        except Exception:
            return nota_id, None

    resumos: dict[int, dict[str, str]] = {}
    workers = max(1, min(16, int(settings.download_storage_workers or 16)))
    with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="relatorio-xml") as executor:
        for nota_id, resumo in executor.map(carregar_xml, tarefas):
            if resumo and resumo.get("tipo_xml") != "evento":
                resumos[nota_id] = resumo
    return resumos


def exportar_conferencia_xlsx(db: Session, filtros: NotasDownloadFiltros) -> tuple[bytes, str]:
    notas = _buscar_notas_relatorio(db, filtros)
    xml_resumos = _xml_resumos_por_nota(db, notas)
    headers = [
        "Competencia",
        "Municipio",
        "Chave de Acesso",
        "Data de Emissao",
        "CNPJ/CPF",
        "Razao Social",
        "No Documento",
        "Valor Total",
        "Valor B/C",
        "CSRF",
        "IRRF",
        "INSS",
        "ISS",
        "Aliquota ISS",
        "Valor Liquido",
        "Valor Liquido Correto",
        "Status Valor Liquido",
        "Campos ausentes no XML",
        "Incidencia ISS",
        "Codigo de servico",
        "Descricao servico nacional",
        "Descricao detalhada do servico",
        "CNAE",
        "Simples Nacional / XML",
        "Status Simples Nacional",
        "Status nota",
        "Divergencia",
        "Prioridade",
        "Responsavel",
        "SLA",
        "Observacao interna",
        "Status CSRF",
        "Status IRRF",
        "Status INSS",
        "Subitem LC116",
        "Codigo Servico Nacional",
        "Origem Base de Calculo",
        "ISS Retido",
        "Valor ISS Retido",
        "PIS Informado",
        "PIS Calculado",
        "COFINS Informado",
        "COFINS Calculado",
        "CSLL Informado",
        "CSLL Calculado",
        "CSRF Calculado",
        "IRRF Calculado",
        "INSS Calculado",
        "ISS Calculado",
        "Status ISS",
        "Valor Liquido Calculado",
        "Regra IRRF",
        "Aliquota Regra IRRF",
        "Regra PIS/COFINS/CSLL",
        "Regra INSS",
        "Observacao Regra",
        "Alertas Fiscais",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Conferencia"
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    even_fill = PatternFill("solid", fgColor="F7FBFF")
    status_ok_fill = PatternFill("solid", fgColor="E2F0D9")
    status_warn_fill = PatternFill("solid", fgColor="FCE4D6")
    border_color = "D9E2F3"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    for nota in notas:
        ws.append([_xlsx_value(value) for value in _nota_relatorio_row(nota, xml_resumos.get(int(nota.id)))])

    money_headers = {
        "Valor Total",
        "Valor B/C",
        "CSRF",
        "IRRF",
        "INSS",
        "ISS",
        "Valor Liquido",
        "Valor Liquido Correto",
        "Valor ISS Retido",
        "PIS Informado",
        "PIS Calculado",
        "COFINS Informado",
        "COFINS Calculado",
        "CSLL Informado",
        "CSLL Calculado",
        "CSRF Calculado",
        "IRRF Calculado",
        "INSS Calculado",
        "ISS Calculado",
        "Valor Liquido Calculado",
    }
    percent_headers = {"Aliquota ISS", "Aliquota Regra IRRF"}
    long_text_headers = {
        "Descricao servico nacional",
        "Descricao detalhada do servico",
        "Observacao interna",
        "Observacao Regra",
        "Alertas Fiscais",
        "Campos ausentes no XML",
    }
    status_headers = {
        "Status Valor Liquido",
        "Status Simples Nacional",
        "Status nota",
        "Divergencia",
        "Status CSRF",
        "Status IRRF",
        "Status INSS",
        "Status ISS",
    }
    header_by_col = {index + 1: header for index, header in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            header = header_by_col.get(cell.column, "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=header in long_text_headers)
            if cell.row % 2 == 0:
                cell.fill = even_fill
            if header in money_headers:
                cell.number_format = '#,##0.00'
            elif header in percent_headers:
                cell.number_format = '0.00'
            if header in status_headers:
                normalized = str(cell.value or "").strip().lower()
                if normalized in {"ok", "correta", "regular", "sem_divergencia"}:
                    cell.fill = status_ok_fill
                elif normalized not in {"", "0"}:
                    cell.fill = status_warn_fill

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        header = column_cells[0].value
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        max_width = 72 if header in long_text_headers else 42
        min_width = 18 if header in long_text_headers else 12
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, min_width), max_width)
    ws.row_dimensions[1].height = 32
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 22

    if ws.max_row >= 1 and ws.max_column >= 1:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName="TabelaConferencia", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        # Excel can repair/remove generated table XML in filtered reports.
        # Keep worksheet AutoFilter and styling without adding an XLSX table.
    output = io.BytesIO()
    wb.save(output)
    filename = f"relatorio_conferencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output.getvalue(), filename


RELATORIO_XLSX_HEADERS = [
    "Competência", "Município", "Data de Emissão", "CNPJ/CPF", "Razão Social", "N° Documento",
    "Valor Total", "Valor B/C", "Retenção CSRF", "CSRF", "IRRF", "Percentual IRRF", "INSS", "ISS",
    "Valor Líquido", "Incidência do ISS", "Data do pagamento", "Código de serviço", "Descrição do Serviço",
    "Código NBS", "Código CNAE", "Descrição CNAE", "Simples Nacional / XML", "Consulta Simples API",
    "Status Simples Nacional", "Status CSRF", "Status IRRF", "Status INSS", "Alertas Fiscais", "dia processado",
]
_RELATORIO_MONEY_HEADERS = {"Valor Total", "Valor B/C", "CSRF", "IRRF", "Percentual IRRF", "INSS", "ISS", "Valor Líquido"}
_RELATORIO_DATE_HEADERS = {"Competência", "Data de Emissão", "Data do pagamento", "dia processado"}
_RELATORIO_WRAP_HEADERS = {"Descrição do Serviço", "Alertas Fiscais"}
_RELATORIO_STATUS_HEADERS = {"Status Simples Nacional", "Status CSRF", "Status IRRF", "Status INSS"}
_INVERTEXTO_CACHE: dict[str, tuple[datetime, dict]] = {}


def _relatorio_digits(value) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def _relatorio_decimal(value) -> Decimal:
    parsed = _parse_decimal_optional(value)
    return parsed if parsed is not None else Decimal("0.00")


def _relatorio_money(value) -> float:
    return float(_relatorio_decimal(value).quantize(Decimal("0.01")))


def _relatorio_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"None", "null", "NaN", "-"} else text


def _relatorio_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _relatorio_text(value)
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError:
        return text[:10]


def _relatorio_codigo_servico(value) -> str:
    digits = _relatorio_digits(value)
    if not digits:
        return ""
    if len(digits) <= 5:
        digits = digits.zfill(6)
    return f"{digits[:2]}.{digits[2:4]}" if len(digits) >= 4 else digits


def _relatorio_status(value, default: str = "Não se aplica") -> str:
    text = _relatorio_text(value)
    normalized = text.lower()
    if not text:
        return default
    if "depend" in normalized or "analise" in normalized or "análise" in normalized:
        return "Depende de análise"
    if "diverg" in normalized or "indevido" in normalized or "nao retido" in normalized or "não retido" in normalized:
        return "Divergente"
    if normalized in {"ok", "correto", "correta", "regular", "sem divergencia", "sem divergência"}:
        return "Correto"
    if "nao se aplica" in normalized or "não se aplica" in normalized or normalized == "n/a":
        return "Não se aplica"
    if "nao comparado" in normalized or "não comparado" in normalized:
        return "Não comparado"
    if "pendente" in normalized:
        return "Pendente"
    return text


def _relatorio_simples(value) -> str:
    text = _relatorio_text(value)
    normalized = text.lower()
    if not text:
        return "Não informado"
    if "mei" in normalized:
        return "MEI"
    if "optante" in normalized and "nao" not in normalized and "não" not in normalized:
        return "Optante S.N"
    if "nao optante" in normalized or "não optante" in normalized:
        return "Não optante"
    return text


def _relatorio_alertas(*values) -> str:
    parts: list[str] = []
    for value in values:
        candidates = value if isinstance(value, (list, tuple, set)) else str(value or "").replace(";", "|").split("|")
        for item in candidates:
            text = _relatorio_text(item)
            if text and text not in parts:
                parts.append(text)
    return " | ".join(parts)


def _relatorio_alerta_critico(alertas: str) -> bool:
    normalized = alertas.lower()
    return any(token in normalized for token in ("diverg", "devido", "não retido", "nao retido", "deveria", "crítico", "critico"))


def _api_flag_bool(value) -> bool | None:
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    normalized = str(value).strip().lower()
    if normalized in {"s", "sim", "true", "1", "yes", "y"}:
        return True
    if normalized in {"n", "nao", "não", "false", "0", "no"}:
        return False
    return None


def _relatorio_status_calculado(informado, calculado, status, tributo: str) -> tuple[str, str]:
    status_text = _relatorio_status(status, default="")
    informado_dec = _relatorio_decimal(informado)
    calculado_dec = _relatorio_decimal(calculado)
    if status_text:
        return status_text, ""
    if calculado_dec == 0 and informado_dec == 0:
        return "Não se aplica", ""
    if informado_dec.quantize(Decimal("0.01")) == calculado_dec.quantize(Decimal("0.01")):
        return "Correto", ""
    return "Divergente", f"{tributo} esperado {calculado_dec:.2f}, informado {informado_dec:.2f}"


def _normalizar_invertexto_payload(payload: dict | None) -> dict:
    if not payload:
        return {"consulta": "Não disponível", "cnae": "", "descricao_cnae": ""}
    simples = payload.get("simples") or {}
    mei = payload.get("mei") or payload.get("simei") or {}
    mei_optante = _api_flag_bool(mei.get("optante") if "optante" in mei else mei.get("optante_mei"))
    simples_optante = _api_flag_bool(
        simples.get("optante") if "optante" in simples else simples.get("optante_simples")
    )
    if mei_optante is True:
        consulta = "MEI"
    elif simples_optante is True:
        consulta = "Optante S.N"
    elif simples_optante is False:
        consulta = "Não optante"
    else:
        consulta = "Não disponível"
    atividade = payload.get("atividade_principal") or payload.get("atividadePrincipal") or {}
    if isinstance(atividade, list):
        atividade = atividade[0] if atividade else {}
    cnae = _relatorio_digits(atividade.get("code") or atividade.get("codigo") or payload.get("cnae"))
    descricao = atividade.get("text") or atividade.get("descricao") or payload.get("descricao_cnae") or ""
    return {"consulta": consulta, "cnae": cnae, "descricao_cnae": _relatorio_text(descricao)}


def _invertexto_result(
    consulta: str,
    cnae: str | None = "",
    descricao_cnae: str | None = "",
) -> dict:
    return {
        "consulta": consulta,
        "consulta_simples_api": consulta,
        "cnae": cnae or "",
        "codigo_cnae": cnae or "",
        "descricao_cnae": descricao_cnae or "",
    }


def _consultar_invertexto_cnpjs(db: Session, cnpjs: set[str]) -> dict[str, dict]:
    cnpjs_normalizados = sorted({cnpj_cache_service.only_digits(cnpj) for cnpj in cnpjs if cnpj_cache_service.only_digits(cnpj)})
    results: dict[str, dict] = {}

    for cnpj in cnpjs_normalizados:
        cache = cnpj_cache_service.get_cache_valido(db, cnpj)
        if cache:
            result = _invertexto_result(
                cache.get("consulta_simples_api") or "Não disponível",
                cache.get("codigo_cnae"),
                cache.get("descricao_cnae"),
            )
            results[cnpj] = result
            _INVERTEXTO_CACHE[cnpj] = (datetime.now(), result)

    pendentes = [cnpj for cnpj in cnpjs_normalizados if cnpj not in results]
    if not pendentes:
        return results

    if not settings.invertexto_enabled:
        results.update({cnpj: _invertexto_result("Não consultado") for cnpj in pendentes})
        return results
    if not settings.invertexto_token:
        results.update({cnpj: _invertexto_result("Não consultado") for cnpj in pendentes})
        return results
    try:
        import requests
    except Exception:
        results.update({cnpj: _invertexto_result("Erro na consulta") for cnpj in pendentes})
        return results

    delay = max(float(settings.invertexto_delay_seconds), 60.0 / max(1, int(settings.invertexto_rpm)))
    for index, cnpj in enumerate(pendentes):
        if index:
            time.sleep(delay)
        try:
            response = requests.get(f"https://api.invertexto.com/v1/cnpj/{cnpj}", params={"token": settings.invertexto_token}, timeout=20)
            response.raise_for_status()
            payload = response.json()
            normalizado = _normalizar_invertexto_payload(payload)
            result = _invertexto_result(normalizado.get("consulta") or "Não disponível", normalizado.get("cnae"), normalizado.get("descricao_cnae"))
            cnpj_cache_service.salvar_cache(
                db,
                cnpj,
                consulta_simples_api=result["consulta"],
                codigo_cnae=result["cnae"],
                descricao_cnae=result["descricao_cnae"],
                status_consulta="OK",
                json_resposta=payload,
            )
            db.commit()
        except Exception as exc:
            db.rollback()
            result = _invertexto_result("Erro na consulta")
            try:
                cnpj_cache_service.salvar_cache(
                    db,
                    cnpj,
                    consulta_simples_api=result["consulta"],
                    codigo_cnae="",
                    descricao_cnae="",
                    status_consulta="Erro na consulta",
                    json_resposta=None,
                    erro=str(exc)[:1000],
                    cache_days=1,
                )
                db.commit()
            except Exception:
                db.rollback()
        _INVERTEXTO_CACHE[cnpj] = (datetime.now(), result)
        results[cnpj] = result
    return results


def _invertexto_cache_cnpjs(db: Session, cnpjs: set[str]) -> dict[str, dict]:
    """Le somente o cache; gerar relatorio nunca deve aguardar API externa."""
    normalizados = {cnpj_cache_service.only_digits(cnpj) for cnpj in cnpjs if cnpj_cache_service.only_digits(cnpj)}
    caches = cnpj_cache_service.get_caches_validos(db, normalizados)
    return {
        cnpj: _invertexto_result(
            caches[cnpj].get("consulta_simples_api") or "Não disponível",
            caches[cnpj].get("codigo_cnae"),
            caches[cnpj].get("descricao_cnae"),
        )
        if cnpj in caches
        else _invertexto_result("Não consultado")
        for cnpj in normalizados
    }


def _relatorio_status_simples(simples_xml: str, consulta_api: str) -> str:
    if consulta_api == "Não consultado":
        return "Não comparado"
    if consulta_api in {"Erro na consulta", "Não disponível"}:
        return "Pendente"
    if simples_xml == "Não informado":
        return "Pendente"
    return "Correto" if simples_xml == consulta_api else "Divergente"


def _relatorio_empresa_cnpjs(db: Session, notas: list[Nota]) -> dict[int, str]:
    empresa_ids = sorted({int(nota.empresa_id) for nota in notas if nota.empresa_id is not None})
    if not empresa_ids:
        return {}
    rows = db.query(Empresa.id, Empresa.cnpj).filter(Empresa.id.in_(empresa_ids)).all()
    return {int(empresa_id): _relatorio_digits(cnpj) for empresa_id, cnpj in rows}


def _relatorio_party(nota: Nota, empresa_cnpj: str) -> tuple[str, str, str]:
    tipo = getattr(notas_service, "_classificar_nota")(nota, empresa_cnpj) if empresa_cnpj else "tomada"
    if tipo == "prestada":
        return _relatorio_digits(getattr(nota, "tomador_cnpj", None)), _relatorio_text(getattr(nota, "tomador_nome", None)), tipo
    return _relatorio_digits(getattr(nota, "prestador_cnpj", None)), _relatorio_text(getattr(nota, "prestador_nome", None)), tipo


def _relatorio_row(nota: Nota, xml_resumo: dict[str, str] | None, empresa_cnpj: str, api_data: dict, gerado_em: date) -> list:
    xml_resumo = xml_resumo or {}
    cnpj, razao_social, _tipo = _relatorio_party(nota, empresa_cnpj)
    simples_xml = _relatorio_simples(_nota_field(nota, "simples_nacional_xml", xml_resumo) or _nota_field(nota, "simples_xml", xml_resumo))
    api_result = api_data.get(cnpj) or {"consulta": "Não consultado", "cnae": "", "descricao_cnae": ""}
    consulta_api = api_result.get("consulta") or "Não consultado"
    valor_total = _relatorio_decimal(_nota_decimal_field(nota, "valor_servico", xml_resumo, "valor_servico"))
    desconto_incondicionado = _relatorio_decimal(xml_resumo.get("valor_desconto_incondicionado"))
    valor_base = _parse_decimal_optional(_nota_decimal_field(nota, "valor_base", xml_resumo, "valor_base"))
    if valor_base is None:
        valor_base = max(Decimal("0"), valor_total - desconto_incondicionado)
    csrf = _relatorio_decimal(_nota_decimal_field(nota, "csrf", xml_resumo, "csrf"))
    irrf = _relatorio_decimal(_nota_decimal_field(nota, "irrf", xml_resumo, "irrf"))
    inss = _relatorio_decimal(_nota_decimal_field(nota, "inss", xml_resumo, "inss"))
    iss_retido = _relatorio_decimal(_nota_decimal_field(nota, "valor_iss_retido", xml_resumo))
    iss_apurado = _relatorio_decimal(_nota_decimal_field(nota, "iss", xml_resumo, "iss"))
    iss = iss_retido if iss_retido > 0 else iss_apurado
    liquido = _parse_decimal_optional(_nota_decimal_field(nota, "valor_liquido", xml_resumo, "valor_liquido"))
    if liquido is None:
        liquido = max(Decimal("0"), valor_total - csrf - irrf - inss - iss_retido)
    status_csrf, alerta_csrf = _relatorio_status_calculado(csrf, _nota_decimal_field(nota, "csrf_calculado", xml_resumo), getattr(nota, "status_csrf", None) or xml_resumo.get("status_csrf"), "CSRF")
    status_irrf, alerta_irrf = _relatorio_status_calculado(irrf, _nota_decimal_field(nota, "irrf_calculado", xml_resumo), getattr(nota, "status_irrf", None) or xml_resumo.get("status_irrf"), "IRRF")
    status_inss, alerta_inss = _relatorio_status_calculado(inss, _nota_decimal_field(nota, "inss_calculado", xml_resumo), getattr(nota, "status_inss", None) or xml_resumo.get("status_inss"), "INSS")
    alertas = _relatorio_alertas(getattr(nota, "alertas_fiscais", None), xml_resumo.get("alertas_fiscais"), alerta_csrf, alerta_irrf, alerta_inss)
    cnae = _relatorio_text(getattr(nota, "cnae", None) or xml_resumo.get("codigo_cnae") or api_result.get("cnae"))
    codigo_servico = _relatorio_codigo_servico(
        _nota_field(nota, "codigo_servico", xml_resumo)
        or _nota_field(nota, "codigo_servico_nacional", xml_resumo)
        or _nota_field(nota, "subitem_lc116", xml_resumo)
    )
    return [
        _relatorio_date(xml_resumo.get("competencia") or nota.competencia or nota.data_emissao),
        _relatorio_text(_nota_field(nota, "municipio", xml_resumo) or _nota_field(nota, "incidencia_iss", xml_resumo)),
        _relatorio_date(xml_resumo.get("data_emissao") or nota.data_emissao),
        cnpj,
        razao_social,
        _relatorio_text(getattr(nota, "numero_nfse", None) or xml_resumo.get("numero_nfse")),
        float(valor_total.quantize(Decimal("0.01"))),
        float(valor_base.quantize(Decimal("0.01"))),
        "1 : Com retenção" if csrf > 0 else "2 : Sem retenção",
        float(csrf.quantize(Decimal("0.01"))),
        float(irrf.quantize(Decimal("0.01"))),
        _relatorio_money(_nota_decimal_field(nota, "regra_irrf_aliquota", xml_resumo)),
        float(inss.quantize(Decimal("0.01"))),
        float(iss.quantize(Decimal("0.01"))),
        float(liquido.quantize(Decimal("0.01"))),
        _relatorio_text(_nota_field(nota, "incidencia_iss", xml_resumo)),
        _relatorio_date(xml_resumo.get("data_pagamento") or getattr(nota, "data_pagamento", None)),
        codigo_servico,
        _relatorio_text(_nota_field(nota, "descricao_servico_detalhada", xml_resumo) or _nota_field(nota, "descricao_servico_nacional", xml_resumo)),
        _relatorio_text(xml_resumo.get("codigo_nbs") or getattr(nota, "codigo_nbs", None) or "0"),
        cnae or "0",
        _relatorio_text(xml_resumo.get("descricao_cnae") or api_result.get("descricao_cnae")),
        simples_xml,
        consulta_api,
        _relatorio_status_simples(simples_xml, consulta_api),
        status_csrf,
        status_irrf,
        status_inss,
        alertas,
        gerado_em,
    ]


def _relatorio_row_divergente(row: list) -> bool:
    by_header = dict(zip(RELATORIO_XLSX_HEADERS, row))
    if any(by_header.get(header) == "Divergente" for header in _RELATORIO_STATUS_HEADERS):
        return True
    return _relatorio_alerta_critico(str(by_header.get("Alertas Fiscais") or ""))


def _add_relatorio_sheet(wb: Workbook, title: str, rows: list[list], table_name: str) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(RELATORIO_XLSX_HEADERS)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    status_ok_fill = PatternFill("solid", fgColor="E2F0D9")
    status_warn_fill = PatternFill("solid", fgColor="FCE4D6")
    thin_border = Border(left=Side(style="thin", color="D9E2F3"), right=Side(style="thin", color="D9E2F3"), top=Side(style="thin", color="D9E2F3"), bottom=Side(style="thin", color="D9E2F3"))
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_by_col = {index + 1: header for index, header in enumerate(RELATORIO_XLSX_HEADERS)}
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            header = header_by_col[cell.column]
            # Estilos por celula sao a parte mais cara do openpyxl em relatorios
            # grandes. Mantemos apenas formatos que carregam informacao fiscal.
            if header in _RELATORIO_WRAP_HEADERS:
                cell.alignment = wrap_alignment
            if header in _RELATORIO_MONEY_HEADERS:
                cell.number_format = '#,##0.00'
            elif header in _RELATORIO_DATE_HEADERS and cell.value:
                cell.number_format = 'yyyy-mm-dd'
            if header in _RELATORIO_STATUS_HEADERS:
                if cell.value == "Correto":
                    cell.fill = status_ok_fill
                elif cell.value == "Divergente":
                    cell.fill = status_warn_fill
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        max_width = 80 if header in _RELATORIO_WRAP_HEADERS else 34
        min_width = 24 if header in _RELATORIO_WRAP_HEADERS else 12
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, min_width), max_width)
    table_ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    # Excel can repair/remove generated table XML in filtered reports.
    # Keep worksheet AutoFilter and styling without adding an XLSX table.


def exportar_conferencia_xlsx(db: Session, filtros: NotasDownloadFiltros) -> tuple[Path, str]:
    notas = [aplicar_campos_operacionais(nota) for nota in _buscar_notas_relatorio(db, filtros)]
    xml_resumos = _xml_resumos_por_nota(db, notas)
    empresa_cnpjs = _relatorio_empresa_cnpjs(db, notas)
    cnpjs_consulta = {cnpj for nota in notas for cnpj, _nome, _tipo in [_relatorio_party(nota, empresa_cnpjs.get(int(nota.empresa_id), ""))] if cnpj}
    api_data = _invertexto_cache_cnpjs(db, cnpjs_consulta)
    gerado_em = datetime.now().date()
    rows = [
        _relatorio_row(nota, xml_resumos.get(int(nota.id)), empresa_cnpjs.get(int(nota.empresa_id), ""), api_data, gerado_em)
        for nota in notas
    ]
    divergentes = [row for row in rows if _relatorio_row_divergente(row)]
    corretas = [row for row in rows if not _relatorio_row_divergente(row)]
    wb = Workbook()
    wb.remove(wb.active)
    _add_relatorio_sheet(wb, "Todas as Notas", rows, "TabelaTodasNotas")
    _add_relatorio_sheet(wb, "Notas Divergentes", divergentes, "TabelaNotasDivergentes")
    _add_relatorio_sheet(wb, "Notas Corretas", corretas, "TabelaNotasCorretas")
    filename = f"relatorio_conferencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    descriptor, temp_name = tempfile.mkstemp(prefix="relatorio_conferencia_", suffix=".xlsx")
    os.close(descriptor)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    return temp_path, filename
