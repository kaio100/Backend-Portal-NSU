from __future__ import annotations

import os
import io
from datetime import date, datetime, timedelta, timezone

from cryptography.fernet import Fernet
from openpyxl import load_workbook

os.environ["DATABASE_URL"] = "sqlite:///./data/test_consultas_api.db"
os.environ["API_WORKER_ENABLED"] = "false"
os.environ["WORKER_DRY_RUN"] = "true"
os.environ["SECRETS_KEY"] = Fernet.generate_key().decode("utf-8")

from fastapi.testclient import TestClient  # noqa: E402

from backend.app.core.config import settings  # noqa: E402
from backend.app.db.models import (  # noqa: E402
    Arquivo,
    Certificado,
    CnpjCache,
    Empresa,
    Evento,
    Job,
    LockProcessamento,
    LogProcesso,
    Nota,
    NsuControle,
    Processo,
)
from backend.app.db.session import SessionLocal, init_db  # noqa: E402
from backend.app.main import app  # noqa: E402
from backend.app.services import cnpj_cache_service, cnpj_enrichment_service, portal_support_service  # noqa: E402
from backend.app.services.storage_service import get_storage_service  # noqa: E402


EXPECTED_RELATORIO_HEADERS = [
    "Competência",
    "Município",
    "Data de Emissão",
    "CNPJ/CPF",
    "Razão Social",
    "N° Documento",
    "Valor Total",
    "Valor B/C",
    "Retenção CSRF",
    "CSRF",
    "IRRF",
    "Percentual IRRF",
    "INSS",
    "ISS",
    "Valor Líquido",
    "Incidência do ISS",
    "Data do pagamento",
    "Código de serviço",
    "Descrição do Serviço",
    "Código NBS",
    "Código CNAE",
    "Descrição CNAE",
    "Simples Nacional / XML",
    "Consulta Simples API",
    "Status Simples Nacional",
    "Status CSRF",
    "Status IRRF",
    "Status INSS",
    "Alertas Fiscais",
    "dia processado",
]


def _reset_db() -> None:
    init_db()
    with SessionLocal() as db:
        db.query(LogProcesso).delete()
        db.query(Job).delete()
        db.query(Evento).delete()
        db.query(Arquivo).delete()
        db.query(Nota).delete()
        db.query(NsuControle).delete()

        db.query(CnpjCache).delete()
        db.query(Processo).delete()
        db.query(LockProcessamento).delete()
        db.query(Certificado).delete()
        db.query(Empresa).delete()
        db.commit()


def _base_data():
    storage = get_storage_service()
    with SessionLocal() as db:
        empresa = Empresa(nome="Guardia Teste LTDA", cnpj="26743708000126", ambiente="producao", ativo=True)
        db.add(empresa)
        db.flush()
        processo = Processo(
            empresa_id=empresa.id,
            certificado_id=None,
            tipo="consulta_nfse",
            status="finalizado",
            nsu_inicio=10,
            nsu_final=20,
        )
        db.add(processo)
        db.flush()
        nota = Nota(
            empresa_id=empresa.id,
            processo_id=processo.id,
            chave="CHAVE-PORTAL-1",
            numero_nfse="123",
            data_emissao=date(2026, 6, 15),
            competencia=date(2026, 6, 1),
            prestador_cnpj="11111111000191",
            prestador_nome="Prestador Portal",
            tomador_cnpj=empresa.cnpj,
            tomador_nome=empresa.nome,
            valor_servico=1000,
            valor_liquido=900,
            valor_base=1000,
            irrf=0,
            irrf_calculado=15,
            status_irrf="divergente",
            status_documento="autorizada",
        )
        db.add(nota)
        db.flush()
        xml_key = f"test-portal/{nota.id}/nota.xml"
        pdf_key = f"test-portal/{nota.id}/nota.pdf"
        storage.put_bytes(xml_key, b"<NFSe/>", content_type="application/xml")
        storage.put_bytes(pdf_key, b"%PDF-1.4", content_type="application/pdf")
        db.add(
            Arquivo(
                empresa_id=empresa.id,
                processo_id=processo.id,
                nota_id=nota.id,
                tipo="XML",
                storage_backend=storage.backend,
                storage_bucket=settings.storage_bucket,
                storage_key=xml_key,
                filename="Prestador Portal NFS-e 123.xml",
                content_type="application/xml",
                tamanho_bytes=7,
            )
        )
        db.add(
            Arquivo(
                empresa_id=empresa.id,
                processo_id=processo.id,
                nota_id=nota.id,
                tipo="PDF_ESPELHO",
                storage_backend=storage.backend,
                storage_bucket=settings.storage_bucket,
                storage_key=pdf_key,
                filename="Prestador Portal NFS-e 123.pdf",
                content_type="application/pdf",
                tamanho_bytes=8,
            )
        )
        db.add(
            Evento(
                empresa_id=empresa.id,
                nota_id=nota.id,
                chave_evento="EVENTO-1",
                chave_afetada=nota.chave,
                tipo_evento="cancelamento",
                descricao="Cancelamento de NFS-e",
                data_evento=datetime(2026, 6, 16, tzinfo=timezone.utc),
                nsu=21,
            )
        )
        db.add(NsuControle(empresa_id=empresa.id, certificado_id=None, cnpj=empresa.cnpj, ultimo_nsu=20))
        db.commit()
        return int(empresa.id), int(processo.id), int(nota.id)


def test_detalhe_documentos_eventos_e_comparativo_da_nota():
    _reset_db()
    _, _, nota_id = _base_data()
    with TestClient(app) as client:
        detalhe = client.get(f"/notas/{nota_id}")
        assert detalhe.status_code == 200
        assert detalhe.json()["empresa_nome"] == "Guardia Teste LTDA"
        assert detalhe.json()["status_nota"] == "autorizada"

        arquivos_antigo = client.get(f"/notas/{nota_id}/arquivos")
        assert arquivos_antigo.status_code == 200
        assert isinstance(arquivos_antigo.json(), list)
        assert {item["tipo"] for item in arquivos_antigo.json()} == {"XML", "PDF_ESPELHO"}

        arquivos = client.get(f"/notas/{nota_id}/arquivos", params={"detalhado": "true"})
        assert arquivos.status_code == 200
        assert arquivos.json()["nota_id"] == nota_id
        assert {item["tipo"] for item in arquivos.json()["items"]} == {"xml", "pdf"}

        eventos = client.get(f"/notas/{nota_id}/eventos")
        assert eventos.status_code == 200
        assert eventos.json()["items"][0]["descricao"] == "Cancelamento de NFS-e"

        comparativo = client.get(f"/notas/{nota_id}/tributos-comparativo")
        assert comparativo.status_code == 200
        assert comparativo.json()["items"][0]["tributo"] == "IRRF"
        assert comparativo.json()["items"][0]["status"] == "divergente"


def test_comparativo_tributos_usa_calculado_correto_e_esconde_observacao_ok():
    _reset_db()
    empresa_id, processo_id, _ = _base_data()
    with SessionLocal() as db:
        empresa = db.get(Empresa, empresa_id)
        nota = Nota(
            empresa_id=empresa.id,
            processo_id=processo_id,
            chave="CHAVE-COMPARATIVO-1",
            numero_nfse="456",
            data_emissao=date(2026, 7, 1),
            competencia=date(2026, 7, 1),
            prestador_cnpj="22222222000182",
            prestador_nome="Prestador Comparativo",
            tomador_cnpj=empresa.cnpj,
            tomador_nome=empresa.nome,
            valor_servico=1000,
            status_documento="autorizada",
            # INSS: informado bate com o calculado -> nao deve aparecer como divergente
            # nem usar o "informado" como se fosse o "calculado" (bug antigo).
            inss=110,
            inss_calculado=110,
            status_inss="Correto",
            # ISS: nao retido -> nao ha calculo/comparacao de aliquota a fazer.
            iss=16,
            iss_calculado=None,
            status_iss="Nao Retido",
            # IRRF: divergente de verdade, deve continuar aparecendo com observacao.
            irrf=0,
            irrf_calculado=15,
            status_irrf="Divergente",
        )
        db.add(nota)
        db.commit()
        nota_id = int(nota.id)

    with TestClient(app) as client:
        response = client.get(f"/notas/{nota_id}/tributos-comparativo")
        assert response.status_code == 200
        items = {item["tributo"]: item for item in response.json()["items"]}

        assert items["INSS"]["calculado"] == 110
        assert items["INSS"]["status"] == "Correto"
        assert items["INSS"]["observacao"] is None

        assert items["ISS"]["status"] == "Nao Retido"
        assert items["ISS"]["observacao"] is None

        assert items["IRRF"]["status"] == "Divergente"
        assert items["IRRF"]["observacao"] == "IRRF esperado diferente do informado"


def test_nota_substituida_sem_evento_explicito_exibe_evento_e_observacao_interna():

    _reset_db()

    _, _, nota_id = _base_data()

    with SessionLocal() as db:

        db.query(Evento).delete()

        nota = db.get(Nota, nota_id)

        nota.status_documento = "substituida"

        nota.status_rotulo = "Substituida"

        db.commit()

    with TestClient(app) as client:

        detalhe = client.get(f"/notas/{nota_id}")

        assert detalhe.status_code == 200

        assert "substituida" in detalhe.json()["observacao_interna"].lower()

        eventos = client.get(f"/notas/{nota_id}/eventos")

        assert eventos.status_code == 200

        assert eventos.json()["total"] == 1

        assert eventos.json()["items"][0]["status"] == "substituida"

        assert "substituida" in eventos.json()["items"][0]["descricao"].lower()




def test_eventos_globais_conferencia_e_busca_avancada():
    _reset_db()
    _, _, nota_id = _base_data()
    with TestClient(app) as client:
        eventos = client.get("/eventos", params={"tipo_evento": "cancelamento"})
        assert eventos.status_code == 200
        assert eventos.json()["total"] == 1

        patch = client.patch(
            f"/notas/{nota_id}/conferencia",
            json={
                "conferencia_status": "corrigir",
                "observacao": "Conferir imposto",
                "responsavel": "Kaio",
                "prioridade": "alta",
                "prioridade_manual": "alta",
                "divergencia": "ISS divergente",
                "valor_liquido_correto": 850,
                "atualizado_por": "Kaio",
            },
        )
        assert patch.status_code == 200
        payload = patch.json()
        assert payload["conferencia_observacao"] == "Conferir imposto"
        assert payload["responsavel"] == "Kaio"
        assert payload["divergencia"] == "ISS divergente"
        assert payload["valor_liquido_correto"] == "850.00"
        assert payload["status_valor_liquido"] == "divergente"

        busca = client.get("/notas", params={"q": "Prestador Portal", "prioridade": "alta"})
        assert busca.status_code == 200
        assert [item["id"] for item in busca.json()] == [nota_id]


def test_endpoints_de_processo_e_resumo_empresa():
    _reset_db()
    empresa_id, processo_id, nota_id = _base_data()
    with TestClient(app) as client:
        arquivos = client.get(f"/processos/{processo_id}/arquivos", params={"tipo": "pdf"})
        assert arquivos.status_code == 200
        assert arquivos.json()["total"] == 1
        assert arquivos.json()["items"][0]["nome"].endswith(".pdf")

        notas = client.get(
            f"/processos/{processo_id}/notas",
            params={"tipo_nota": "recebida", "busca": "Prestador", "valor_min": "999", "somente_divergentes": "true"},
        )
        assert notas.status_code == 200
        assert notas.json()["total"] == 1
        assert notas.json()["items"][0]["id"] == nota_id

        summary = client.get(f"/processos/{processo_id}/summary")
        assert summary.status_code == 200
        assert summary.json()["total_notas"] == 1
        assert summary.json()["total_xml"] == 1
        assert summary.json()["total_pdf"] == 1
        assert summary.json()["nsu_final"] == 20

        resumo = client.get("/empresas/resumo-operacional")
        assert resumo.status_code == 200
        item = next(item for item in resumo.json()["items"] if item["empresa_id"] == empresa_id)
        assert item["total_notas"] == 1
        assert item["ultimo_nsu"] == 20


def test_relatorio_conferencia_csv(monkeypatch):
    _reset_db()
    empresa_id, _, _ = _base_data()
    monkeypatch.setattr(
        portal_support_service,
        "_consultar_invertexto_cnpjs",
        lambda db, cnpjs: {cnpj: {"consulta": "Não consultado", "cnae": "", "descricao_cnae": ""} for cnpj in cnpjs},
    )
    with SessionLocal() as db:
        empresa = db.get(Empresa, empresa_id)
        correta = Nota(
            empresa_id=empresa_id,
            processo_id=None,
            chave="CHAVE-PORTAL-2",
            numero_nfse="456",
            data_emissao=date(2026, 6, 20),
            competencia=date(2026, 6, 1),
            prestador_cnpj=empresa.cnpj,
            prestador_nome=empresa.nome,
            tomador_cnpj="22222222000182",
            tomador_nome="Tomador Prestada",
            valor_servico=None,
            valor_base=None,
            irrf=None,
            inss=None,
            csrf=None,
            iss=None,
            valor_liquido=None,
            codigo_servico="010500",
            simples_nacional_xml=None,
            status_csrf="Não se aplica",
            status_irrf="Correto",
            status_inss="Depende de análise",
        )
        db.add(correta)
        db.commit()
    with TestClient(app) as client:
        response = client.post("/relatorios/conferencia", json={"filtros": {"empresa_id": empresa_id}})
        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response.headers["content-disposition"]
        wb = load_workbook(filename=io.BytesIO(response.content))
        assert wb.sheetnames == ["Todas as Notas", "Notas Divergentes", "Notas Corretas"]
        for ws in wb.worksheets:
            headers = [cell.value for cell in ws[1]]
            assert headers == EXPECTED_RELATORIO_HEADERS
            assert ws.freeze_panes == "A2"
            assert ws.auto_filter.ref is not None

        todas = wb["Todas as Notas"]
        divergentes = wb["Notas Divergentes"]
        corretas = wb["Notas Corretas"]
        rows = list(todas.iter_rows(min_row=2, values_only=True))
        row_by_doc = {row[5]: row for row in rows}

        assert len(rows) == 2
        assert divergentes.max_row == 2
        assert corretas.max_row == 2
        assert divergentes["F2"].value == "123"
        assert corretas["F2"].value == "456"
        assert row_by_doc["456"][6] == 0
        assert row_by_doc["456"][7] == 0
        assert row_by_doc["456"][9] == 0
        assert row_by_doc["456"][10] == 0
        assert row_by_doc["456"][11] == 0
        assert row_by_doc["456"][12] == 0
        assert row_by_doc["456"][13] == 0
        assert row_by_doc["456"][14] == 0
        assert row_by_doc["456"][17] == "01.05"
        assert row_by_doc["456"][23] == "Não consultado"
        assert row_by_doc["456"][24] == "Não comparado"
        assert row_by_doc["456"][25] == "Não se aplica"
        assert row_by_doc["456"][27] == "Depende de análise"


def test_relatorio_conferencia_enriquece_campos_vazios_pelo_xml(monkeypatch):
    monkeypatch.setattr(portal_support_service.settings, "invertexto_enabled", False)
    _reset_db()
    empresa_id, _, nota_id = _base_data()
    storage = get_storage_service()
    xml = b"""
<NFSe>
  <infNFSe Id="NFS53001081237381902000125000000000012326010112345678">
    <xLocPrestacao>Brasilia</xLocPrestacao>
    <xLocIncid>Goiania</xLocIncid>
    <xTribNac>Servicos administrativos</xTribNac>
    <DPS>
      <infDPS>
        <serv><cServ><cTribNac>170601</cTribNac><xDescServ>Descricao detalhada do XML</xDescServ></cServ></serv>
        <valores><trib><tribFed><vRetIRRF>10.00</vRetIRRF><vRetCP>20.00</vRetCP><vRetCSLL>3.00</vRetCSLL><piscofins><vPis>1.00</vPis><vCofins>2.00</vCofins></piscofins></tribFed></trib></valores>
      </infDPS>
    </DPS>
    <valores><vBC>1000.00</vBC><vISSQN>50.00</vISSQN><pAliqAplic>5.00</pAliqAplic><vLiq>900.00</vLiq></valores>
  </infNFSe>
</NFSe>
"""
    with SessionLocal() as db:
        nota = db.get(Nota, nota_id)
        nota.municipio = None
        nota.valor_base = None
        nota.csrf = None
        nota.irrf = None
        nota.inss = None
        nota.iss = None
        nota.aliquota_iss = None
        nota.valor_liquido_correto = None
        nota.status_valor_liquido = None
        nota.incidencia_iss = None
        nota.codigo_servico = None
        nota.descricao_servico_nacional = None
        nota.descricao_servico_detalhada = None
        arquivo = db.query(Arquivo).filter(Arquivo.nota_id == nota_id, Arquivo.tipo == "XML").first()
        storage.put_bytes(arquivo.storage_key, xml, content_type="application/xml")
        arquivo.tamanho_bytes = len(xml)
        db.commit()

    with TestClient(app) as client:
        response = client.post("/relatorios/conferencia", json={"filtros": {"empresa_id": empresa_id}})
        assert response.status_code == 200
        wb = load_workbook(filename=io.BytesIO(response.content), read_only=True, data_only=True)
        ws = wb["Todas as Notas"]
        headers = [cell.value for cell in ws[1]]
        row = {header: ws.cell(row=2, column=index + 1).value for index, header in enumerate(headers)}

        assert headers == EXPECTED_RELATORIO_HEADERS
        assert row["Município"] == "Brasilia"
        assert row["Incidência do ISS"] == "Goiania"
        assert row["Código de serviço"] == "17.06"
        assert row["Descrição do Serviço"] == "Descricao detalhada do XML"
        assert row["Valor B/C"] == 1000
        assert row["ISS"] == 50
        assert row["CSRF"] == 6
        assert row["IRRF"] == 10
        assert row["INSS"] == 20
        assert row["Valor Líquido"] == 900
        assert row["Consulta Simples API"] == "Não consultado"


def test_relatorio_codigo_servico_padrao_xx_xx():
    assert portal_support_service._relatorio_codigo_servico("170101") == "17.01"
    assert portal_support_service._relatorio_codigo_servico("110101") == "11.01"
    assert portal_support_service._relatorio_codigo_servico("010500") == "01.05"
    assert portal_support_service._relatorio_codigo_servico("10500") == "01.05"


def test_relatorio_invertexto_usa_cnpjs_unicos_e_cache(monkeypatch):
    _reset_db()
    calls = []

    class FakeResponse:
        def raise_for_status(self):
            return None

        def json(self):
            return {
                "simples": {"optante": True},
                "simei": {"optante": False},
                "atividade_principal": {"code": "6201501", "text": "Desenvolvimento de programas"},
            }

    def fake_get(url, params, timeout):
        calls.append((url, params, timeout))
        return FakeResponse()

    import requests

    monkeypatch.setattr(portal_support_service.settings, "invertexto_enabled", True)
    monkeypatch.setattr(portal_support_service.settings, "invertexto_token", "token-teste")
    monkeypatch.setattr(portal_support_service.settings, "invertexto_delay_seconds", 0)
    monkeypatch.setattr(portal_support_service.settings, "invertexto_rpm", 100000)
    monkeypatch.setattr(requests, "get", fake_get)
    portal_support_service._INVERTEXTO_CACHE.clear()

    with SessionLocal() as db:
        first = portal_support_service._consultar_invertexto_cnpjs(db, {"11222333000181", "11222333000181"})
        second = portal_support_service._consultar_invertexto_cnpjs(db, {"11222333000181"})

    assert len(calls) == 1
    assert first["11222333000181"]["consulta"] == "Optante S.N"
    assert first["11222333000181"]["cnae"] == "6201501"
    assert second["11222333000181"]["descricao_cnae"] == "Desenvolvimento de programas"


def test_relatorio_invertexto_normaliza_formato_atual():
    result = portal_support_service._normalizar_invertexto_payload(
        {
            "simples": {"optante_simples": "N"},
            "mei": {"optante_mei": "N"},
            "atividade_principal": {
                "codigo": "6201501",
                "descricao": "Desenvolvimento de programas de computador sob encomenda",
            },
        }
    )

    assert result["consulta"] == "Não optante"
    assert result["cnae"] == "6201501"
    assert result["descricao_cnae"] == "Desenvolvimento de programas de computador sob encomenda"


def test_relatorio_invertexto_normaliza_mei_formato_atual():
    result = portal_support_service._normalizar_invertexto_payload(
        {
            "simples": {"optante_simples": "S"},
            "mei": {"optante_mei": "S"},
            "atividade_principal": {"codigo": "4711302", "descricao": "Comercio varejista"},
        }
    )

    assert result["consulta"] == "MEI"
    assert result["cnae"] == "4711302"


def test_invertexto_cache_persistente_valido_nao_chama_api(monkeypatch):
    _reset_db()
    portal_support_service._INVERTEXTO_CACHE.clear()
    with SessionLocal() as db:
        cnpj_cache_service.salvar_cache(
            db,
            "11.222.333/0001-81",
            consulta_simples_api="Optante S.N",
            codigo_cnae="6201501",
            descricao_cnae="Desenvolvimento de programas",
            status_consulta="OK",
            json_resposta={"origem": "cache"},
        )
        db.commit()

    calls = []

    def fake_get(url, params, timeout):
        calls.append((url, params, timeout))
        raise AssertionError("API Invertexto nao deveria ser chamada com cache valido")

    import requests

    monkeypatch.setattr(portal_support_service.settings, "invertexto_enabled", True)
    monkeypatch.setattr(portal_support_service.settings, "invertexto_token", "token-teste")
    monkeypatch.setattr(requests, "get", fake_get)

    with SessionLocal() as db:
        result = portal_support_service._consultar_invertexto_cnpjs(db, {"11222333000181"})

    assert calls == []
    assert result["11222333000181"]["consulta"] == "Optante S.N"
    assert result["11222333000181"]["cnae"] == "6201501"
    assert result["11222333000181"]["descricao_cnae"] == "Desenvolvimento de programas"


def test_invertexto_cache_expirado_chama_api_e_atualiza(monkeypatch):
    _reset_db()
    portal_support_service._INVERTEXTO_CACHE.clear()
    with SessionLocal() as db:
        cnpj_cache_service.salvar_cache(
            db,
            "11222333000181",
            consulta_simples_api="Não optante",
            codigo_cnae="0000000",
            descricao_cnae="Antigo",
            status_consulta="OK",
            json_resposta={"origem": "expirado"},
        )
        cache = db.query(CnpjCache).filter(CnpjCache.cnpj == "11222333000181").first()
        cache.data_expiracao = date.today() - timedelta(days=1)
        db.commit()

    calls = []

    class FakeResponse:
        def raise_for_status(self):
            return None

        def json(self):
            return {
                "simples": {"optante_simples": "S"},
                "mei": {"optante_mei": "N"},
                "atividade_principal": {"codigo": "6201501", "descricao": "Novo CNAE"},
            }

    def fake_get(url, params, timeout):
        calls.append((url, params, timeout))
        return FakeResponse()

    import requests

    monkeypatch.setattr(portal_support_service.settings, "invertexto_enabled", True)
    monkeypatch.setattr(portal_support_service.settings, "invertexto_token", "token-teste")
    monkeypatch.setattr(portal_support_service.settings, "invertexto_delay_seconds", 0)
    monkeypatch.setattr(portal_support_service.settings, "invertexto_rpm", 100000)
    monkeypatch.setattr(requests, "get", fake_get)

    with SessionLocal() as db:
        result = portal_support_service._consultar_invertexto_cnpjs(db, {"11222333000181"})
        atualizado = db.query(CnpjCache).filter(CnpjCache.cnpj == "11222333000181").first()

    assert len(calls) == 1
    assert result["11222333000181"]["consulta"] == "Optante S.N"
    assert atualizado.codigo_cnae == "6201501"
    assert atualizado.descricao_cnae == "Novo CNAE"


def test_invertexto_desligado_sem_cache_nao_chama_api(monkeypatch):
    _reset_db()
    portal_support_service._INVERTEXTO_CACHE.clear()
    calls = []

    def fake_get(url, params, timeout):
        calls.append((url, params, timeout))
        raise AssertionError("API Invertexto nao deveria ser chamada desligada")

    import requests

    monkeypatch.setattr(portal_support_service.settings, "invertexto_enabled", False)
    monkeypatch.setattr(portal_support_service.settings, "invertexto_token", "token-teste")
    monkeypatch.setattr(requests, "get", fake_get)

    with SessionLocal() as db:
        result = portal_support_service._consultar_invertexto_cnpjs(db, {"11222333000181"})

    assert calls == []
    assert result["11222333000181"]["consulta"] == "Não consultado"


def test_invertexto_token_vazio_nao_chama_api(monkeypatch):
    _reset_db()
    portal_support_service._INVERTEXTO_CACHE.clear()
    calls = []

    def fake_get(url, params, timeout):
        calls.append((url, params, timeout))
        raise AssertionError("API Invertexto nao deveria ser chamada sem token")

    import requests

    monkeypatch.setattr(portal_support_service.settings, "invertexto_enabled", True)
    monkeypatch.setattr(portal_support_service.settings, "invertexto_token", "")
    monkeypatch.setattr(requests, "get", fake_get)

    with SessionLocal() as db:
        result = portal_support_service._consultar_invertexto_cnpjs(db, {"11222333000181"})

    assert calls == []
    assert result["11222333000181"]["consulta"] == "Não consultado"


def test_invertexto_cache_expira_em_trinta_dias(monkeypatch):
    _reset_db()
    monkeypatch.setattr(cnpj_cache_service.settings, "invertexto_cache_days", 30)

    with SessionLocal() as db:
        cnpj_cache_service.salvar_cache(
            db,
            "11222333000181",
            consulta_simples_api="Optante S.N",
            codigo_cnae="6201501",
            descricao_cnae="Desenvolvimento de programas",
            status_consulta="OK",
            json_resposta={"ok": True},
        )
        db.commit()
        cache = db.query(CnpjCache).filter(CnpjCache.cnpj == "11222333000181").first()

    assert cache.data_expiracao == cache.data_consulta + timedelta(days=30)


def test_enriquecimento_pos_certificado_consulta_cnpjs_do_processo(monkeypatch):
    _reset_db()
    _, processo_id, _ = _base_data()
    chamadas = []

    def fake_consultar(db, cnpjs):
        chamadas.append(set(cnpjs))
        return {
            cnpj: {
                "consulta": "Optante S.N",
                "consulta_simples_api": "Optante S.N",
                "cnae": "6201501",
                "codigo_cnae": "6201501",
                "descricao_cnae": "Desenvolvimento de programas",
            }
            for cnpj in cnpjs
        }

    monkeypatch.setattr(portal_support_service, "_consultar_invertexto_cnpjs", fake_consultar)

    with SessionLocal() as db:
        resumo = cnpj_enrichment_service.enriquecer_cnpjs_do_processo(db, processo_id=processo_id, certificado_id=123)

    assert chamadas == [{"11111111000191"}]
    assert resumo["cnpjs_total"] == 1
    assert resumo["pendentes"] == 1
