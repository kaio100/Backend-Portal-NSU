from __future__ import annotations

import io
import os
import zipfile
from datetime import date

from cryptography.fernet import Fernet
from openpyxl import load_workbook

os.environ["DATABASE_URL"] = "sqlite:///./data/test_consultas_api.db"
os.environ["API_WORKER_ENABLED"] = "false"
os.environ["WORKER_DRY_RUN"] = "true"
os.environ["SECRETS_KEY"] = Fernet.generate_key().decode("utf-8")

from fastapi.testclient import TestClient  # noqa: E402

from backend.app.core.config import settings  # noqa: E402
from backend.app.db.models import Arquivo, Certificado, Empresa, Evento, Job, LockProcessamento, LogProcesso, Nota, Processo  # noqa: E402
from backend.app.db.session import SessionLocal, init_db  # noqa: E402
from backend.app.main import app  # noqa: E402
from backend.app.services.storage_service import get_storage_service  # noqa: E402


def _reset_db() -> None:
    init_db()
    with SessionLocal() as db:
        db.query(LogProcesso).delete()
        db.query(Job).delete()
        db.query(Evento).delete()
        db.query(Arquivo).delete()
        db.query(Nota).delete()
        db.query(Processo).delete()
        db.query(LockProcessamento).delete()
        db.query(Certificado).delete()
        db.query(Empresa).delete()
        db.commit()


def _seed_notas_tipo_direcao() -> tuple[int, int]:
    storage = get_storage_service()
    with SessionLocal() as db:
        empresa = Empresa(nome="Empresa Referencia", cnpj="26743708000126", ambiente="producao", ativo=True)
        db.add(empresa)
        db.flush()
        processo = Processo(empresa_id=empresa.id, tipo="consulta_nfse", status="finalizado")
        db.add(processo)
        db.flush()

        tomada = Nota(
            empresa_id=empresa.id,
            processo_id=processo.id,
            chave="CHAVE-TOMADA",
            numero_nfse="10",
            data_emissao=date(2026, 6, 1),
            competencia=date(2026, 6, 1),
            prestador_cnpj="11111111000191",
            prestador_nome="Prestador Tomada",
            tomador_cnpj=empresa.cnpj,
            tomador_nome=empresa.nome,
            valor_servico=100,
            valor_liquido=100,
            status_documento="autorizada",
        )
        prestada = Nota(
            empresa_id=empresa.id,
            processo_id=processo.id,
            chave="CHAVE-PRESTADA",
            numero_nfse="20",
            data_emissao=date(2026, 6, 2),
            competencia=date(2026, 6, 1),
            prestador_cnpj=empresa.cnpj,
            prestador_nome=empresa.nome,
            tomador_cnpj="22222222000182",
            tomador_nome="Tomador Prestada",
            valor_servico=200,
            valor_liquido=200,
            status_documento="autorizada",
        )
        db.add_all([tomada, prestada])
        db.flush()

        for nota, label in ((tomada, "tomada"), (prestada, "prestada")):
            key = f"test-tipo-direcao/{nota.id}/{label}.xml"
            storage.put_bytes(key, f"<xml>{label}</xml>".encode("utf-8"), content_type="application/xml")
            db.add(
                Arquivo(
                    empresa_id=empresa.id,
                    processo_id=processo.id,
                    nota_id=nota.id,
                    tipo="XML",
                    storage_backend=storage.backend,
                    storage_bucket=settings.storage_bucket,
                    storage_key=key,
                    filename=f"{label}.xml",
                    content_type="application/xml",
                    tamanho_bytes=20,
                )
            )
        db.commit()
        return int(empresa.id), int(processo.id)


def _items(response):
    data = response.json()
    return data.get("items") if isinstance(data, dict) else data


def test_get_notas_filtra_tipo_e_direcao():
    _reset_db()
    _seed_notas_tipo_direcao()

    with TestClient(app) as client:
        todas = _items(client.get("/notas"))
        tomadas = _items(client.get("/notas?tipo_nota=tomada"))
        prestadas = _items(client.get("/notas?tipo_nota=prestada"))
        recebidas = _items(client.get("/notas?direcao_nota=recebida"))
        emitidas = _items(client.get("/notas?direcao_nota=emitida"))

    assert {item["tipo_nota"] for item in todas} == {"tomada", "prestada"}
    assert {item["direcao_nota"] for item in todas} == {"recebida", "emitida"}
    assert [item["tipo_nota"] for item in tomadas] == ["tomada"]
    assert [item["tipo_nota"] for item in prestadas] == ["prestada"]
    assert [item["direcao_nota"] for item in recebidas] == ["recebida"]
    assert [item["direcao_nota"] for item in emitidas] == ["emitida"]
    assert tomadas[0]["cnpj_tomador"] == "26743708000126"
    assert prestadas[0]["cnpj_prestador"] == "26743708000126"


def test_resumo_e_processo_respeitam_tipo_nota():
    _reset_db()
    _, processo_id = _seed_notas_tipo_direcao()

    with TestClient(app) as client:
        resumo_tomadas = client.get("/notas/resumo?tipo_nota=tomada")
        resumo_prestadas = client.get("/notas/resumo?tipo_nota=prestada")
        processo_tomadas = client.get(f"/processos/{processo_id}/notas?tipo_nota=tomada")
        processo_prestadas = client.get(f"/processos/{processo_id}/notas?direcao_nota=emitida")

    assert resumo_tomadas.status_code == 200
    assert resumo_tomadas.json()["total"] == 1
    assert resumo_prestadas.json()["total"] == 1
    assert processo_tomadas.json()["total"] == 1
    assert processo_tomadas.json()["items"][0]["tipo_nota"] == "tomada"
    assert processo_prestadas.json()["total"] == 1
    assert processo_prestadas.json()["items"][0]["tipo_nota"] == "prestada"


def test_relatorio_conferencia_e_download_lote_respeitam_tipo_nota():
    _reset_db()
    _seed_notas_tipo_direcao()

    with TestClient(app) as client:
        relatorio = client.post("/relatorios/conferencia", json={"tipo_nota": "tomada"})
        download = client.post(
            "/notas/download-lote",
            json={"filtros": {"tipo_nota": "prestada"}, "incluir_xml": True, "incluir_pdf": False},
        )

    assert relatorio.status_code == 200
    workbook = load_workbook(io.BytesIO(relatorio.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    flat = "\n".join(str(cell) for row in rows for cell in row if cell is not None)
    assert "Prestador Tomada" in flat
    assert "Tomador Prestada" not in flat

    assert download.status_code == 200
    with zipfile.ZipFile(io.BytesIO(download.content)) as zf:
        payload = "\n".join(zf.read(name).decode("utf-8") for name in zf.namelist())
    assert "<xml>prestada</xml>" in payload
    assert "<xml>tomada</xml>" not in payload
